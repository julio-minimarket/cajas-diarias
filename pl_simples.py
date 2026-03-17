"""
Módulo P&L Simples - Profit & Loss Simple (v2.0 con persistencia)
Genera informes mensuales de Ingresos vs. Egresos por sucursal
Incluye análisis de composición del gasto, estadísticas comparativas y evolución histórica

Autor: Julio Becker
Fecha: Enero 2026
Integrado a: cajas_diarias.py
Versión: 2.0 - Agregado: Persistencia en DB y evolución histórica
"""

import streamlit as st
import pandas as pd
from datetime import datetime, date
import os
from pathlib import Path
import calendar
import io

# ReportLab para generación de PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.platypus import KeepTogether



# Forzar limpieza de caché al inicio
if 'cache_cleared' not in st.session_state:
    st.cache_resource.clear()
    st.cache_data.clear()
    st.session_state['cache_cleared'] = True

# ==================== FUNCIONES DE PROCESAMIENTO DE GASTOS ====================

# ============================================================
# MAPEO MANUAL: CSV -> SUCURSALES
# ============================================================

# Mapeo hardcoded (usado si no hay tabla mapeo_sucursales_csv)
MAPEO_CSV_HARDCODED = {
    'Belfast S.A.': 4,
    'Chicken Chill Minimarket S.A.': 8,
    'Costumbres Argentinas-Minimarket S.A.': 10,
    'Destileria Open 24 S.A.': 13,
    'Friends S.A.S. Patagonia': 5,
    'Liverpool Minimarket S.A.': 7,
    'Minimarket S.A. ': 1,  # Con espacio al final
    'Minimarket S.A.': 1,
    'Open 24 Minimarket S.A.': 11,
    'Pocillos S.A.S. 1885 ': 3,  # Con espacio al final
    'Pocillos S.A.S. 1885': 3,
    'Pocillos S.A.S.Napoles Rgl': 2,
    'Rincon Chico S.A.S. Napoles Calafate': 6,
    'Temple Minimarket S.A.': 9,
    # NOTA: "Andes Food S.A." no mapeada - no existe sucursal
}

def obtener_mapeo_manual(supabase):
    """
    Obtiene mapeo manual desde tabla o hardcoded
    
    Prioridad:
    1. Tabla mapeo_sucursales_csv en Supabase (si existe)
    2. MAPEO_CSV_HARDCODED (fallback)
    
    Returns:
        dict: {nombre_csv: sucursal_id}
    """
    mapeo = {}
    
    # Intentar obtener de la tabla
    try:
        result = supabase.table("mapeo_sucursales_csv")\
            .select("nombre_csv, sucursal_id")\
            .eq("activo", True)\
            .execute()
        
        if result.data:
            # Usar mapeo de la tabla
            for row in result.data:
                # Mapeo exacto
                mapeo[row['nombre_csv']] = row['sucursal_id']
                
                # Mapeo normalizado
                nombre_norm = row['nombre_csv'].upper().replace(' ', '').replace('.', '').replace(',', '')
                mapeo[nombre_norm] = row['sucursal_id']
            
            return mapeo
    except Exception as e:
        # Tabla no existe o error al consultar
        print(f"[INFO] No se pudo obtener mapeo de DB, usando hardcoded: {e}")
    
    # Usar mapeo hardcoded si no hay tabla o está vacía
    if not mapeo:
        mapeo = MAPEO_CSV_HARDCODED.copy()
        
        # Agregar versiones normalizadas
        for nombre_csv, sucursal_id in MAPEO_CSV_HARDCODED.items():
            nombre_norm = nombre_csv.upper().replace(' ', '').replace('.', '').replace(',', '')
            mapeo[nombre_norm] = sucursal_id
    
    return mapeo


# ============================================================
# FUNCIONES DE MAPEO
# ============================================================

def convertir_importe(valor):
    """
    Convierte importes en diferentes formatos a float
    Maneja formato argentino ($1.234.567,89) y formato numérico (1234567.89)
    """
    valor = str(valor).strip()
    
    # Si tiene $ y puntos (formato argentino: $1.234.567,89)
    if '$' in valor:
        valor = valor.replace('$', '').strip()
        valor = valor.replace('.', '')  # Elimino separadores de miles
        valor = valor.replace(',', '.')  # Coma decimal a punto
    # Si tiene coma pero no $ (formato: 1.234.567,89)
    elif ',' in valor:
        valor = valor.replace('.', '')  # Elimino separadores de miles
        valor = valor.replace(',', '.')  # Coma decimal a punto
    
    try:
        return float(valor)
    except:
        return 0.0


def procesar_archivo_gastos(archivo_csv):
    """
    Procesa un archivo CSV de gastos/facturas y retorna un DataFrame con los datos calculados
    Maneja múltiples formatos de fecha
    """
    try:
        # Cargo el archivo
        df = pd.read_csv(archivo_csv)
        
        # Elimino la última fila (totales del mes) si existe
        if len(df) > 0 and pd.isna(df.iloc[-1]['Empresa']):
            df = df[:-1]
        
        # Columnas para calcular NETO
        columnas_neto = ['Importe Neto 21', 'Importe Neto 10_5', 'Importe Neto 27', 'Impuestos Internos']
        
        # Columnas para calcular IVA
        columnas_iva = ['Percepcion Iva', 'Otras Percepciones', 'Percepcion IIBB', 
                        'Iva 21', 'Iva 10,5', 'Iva 27']
        
        # Convierto todas las columnas necesarias
        for col in columnas_neto + columnas_iva:
            if col in df.columns:
                df[col] = df[col].apply(convertir_importe)
        
        # Convertir fechas - Priorizar Fecha C. (Contabilización) sobre Fecha E. (Emisión)
        # Intentar primero con Fecha C.
        if 'Fecha C.' in df.columns:
            df['Fecha'] = pd.to_datetime(df['Fecha C.'], format='%d/%m/%Y', errors='coerce')
        
        # Si no hay Fecha C. o hay NaN, usar Fecha E.
        if 'Fecha E.' in df.columns:
            if 'Fecha' not in df.columns:
                df['Fecha'] = pd.to_datetime(df['Fecha E.'], format='%d/%m/%Y', errors='coerce')
            else:
                # Rellenar NaN de Fecha con Fecha E.
                mask = df['Fecha'].isna()
                df.loc[mask, 'Fecha'] = pd.to_datetime(df.loc[mask, 'Fecha E.'], format='%d/%m/%Y', errors='coerce')
        
        # Calculo NETO e IVA para cada fila
        df['NETO'] = df[columnas_neto].sum(axis=1)
        df['IVA_PERCEPCIONES'] = df[columnas_iva].sum(axis=1)
        df['TOTAL_GASTO'] = df['NETO'] + df['IVA_PERCEPCIONES']
        
        # Elimino filas sin datos válidos
        df = df.dropna(subset=['Empresa'])
        df = df[df['TOTAL_GASTO'] > 0]
        
        return df
        
    except Exception as e:
        st.error(f"❌ Error procesando archivo de gastos: {str(e)}")
        return None


def verificar_gastos_existentes(supabase, sucursal_id, mes, anio):
    """
    Verifica si ya existen gastos cargados para una sucursal en un período específico
    
    Parámetros:
    -----------
    sucursal_id : int o None
        ID de la sucursal. Si es None, verifica para todas las sucursales
    
    Retorna:
    --------
    dict con 'existe': bool, 'cantidad': int, 'total': float
    """
    try:
        query = supabase.table("gastos_mensuales").select("*")
        
        if sucursal_id is not None:
            query = query.eq("sucursal_id", sucursal_id)
        
        query = query.eq("mes", mes).eq("anio", anio)
        
        result = query.execute()
        
        if result.data and len(result.data) > 0:
            total = sum(r['total'] for r in result.data)
            return {
                'existe': True,
                'cantidad': len(result.data),
                'total': total,
                'fecha_importacion': result.data[0].get('fecha_importacion')
            }
        else:
            return {'existe': False, 'cantidad': 0, 'total': 0}
            
    except Exception as e:
        st.error(f"❌ Error verificando gastos existentes: {str(e)}")
        return {'existe': False, 'cantidad': 0, 'total': 0}


def crear_mapeo_sucursales(supabase):
    """
    Crea un diccionario de mapeo: nombre_empresa -> sucursal_id
    Soporta coincidencias parciales y variaciones de nombres
    """
    try:
        result = supabase.table("sucursales").select("id, nombre").execute()
        if not result.data:
            return {}
        
        mapeo = {}
        for sucursal in result.data:
            nombre = sucursal['nombre']
            sucursal_id = sucursal['id']
            
            # Mapeo exacto
            mapeo[nombre] = sucursal_id
            
            # Mapeo normalizado (sin espacios, mayúsculas, puntos)
            nombre_norm = nombre.upper().replace(' ', '').replace('.', '').replace(',', '')
            mapeo[nombre_norm] = sucursal_id
            
            # Mapeo por palabras clave (primeras palabras significativas)
            palabras = nombre.upper().split()
            if len(palabras) > 0:
                # Primera palabra
                mapeo[palabras[0]] = sucursal_id
                # Primeras dos palabras
                if len(palabras) > 1:
                    mapeo[f"{palabras[0]} {palabras[1]}"] = sucursal_id
        
        return mapeo
    except Exception as e:
        st.error(f"❌ Error creando mapeo de sucursales: {str(e)}")
        return {}


def obtener_sucursal_id_desde_nombre(nombre_empresa, mapeo_manual, mapeo_automatico):
    """
    Obtiene el sucursal_id desde el nombre de la empresa del CSV
    
    Prioridades:
    1. Mapeo manual EXACTO (desde tabla o hardcoded)
    2. Mapeo manual NORMALIZADO
    3. Mapeo automático (4 estrategias)
    
    Args:
        nombre_empresa (str): Nombre de la empresa del CSV
        mapeo_manual (dict): Mapeo manual {nombre: id}
        mapeo_automatico (dict): Mapeo automático {nombre: id}
    
    Returns:
        int or None: sucursal_id o None si no se encuentra
    """
    if not nombre_empresa or pd.isna(nombre_empresa):
        return None
    
    nombre_empresa = str(nombre_empresa).strip()
    
    # PRIORIDAD 1: Mapeo manual EXACTO
    if nombre_empresa in mapeo_manual:
        return mapeo_manual[nombre_empresa]
    
    # PRIORIDAD 2: Mapeo manual NORMALIZADO
    nombre_norm = nombre_empresa.upper().replace(' ', '').replace('.', '').replace(',', '')
    if nombre_norm in mapeo_manual:
        return mapeo_manual[nombre_norm]
    
    # PRIORIDAD 3: Mapeo automático
    # Estrategia 1: Coincidencia exacta
    if nombre_empresa in mapeo_automatico:
        return mapeo_automatico[nombre_empresa]
    
    # Estrategia 2: Coincidencia normalizada
    if nombre_norm in mapeo_automatico:
        return mapeo_automatico[nombre_norm]
    
    # Estrategia 3: Por palabras clave
    palabras = nombre_empresa.upper().split()
    if len(palabras) > 0:
        # Primera palabra
        if palabras[0] in mapeo_automatico:
            return mapeo_automatico[palabras[0]]
        # Primeras dos palabras
        if len(palabras) > 1:
            clave = f"{palabras[0]} {palabras[1]}"
            if clave in mapeo_automatico:
                return mapeo_automatico[clave]
    
    # Estrategia 4: Búsqueda parcial (contiene)
    for nombre_mapeado, suc_id in mapeo_automatico.items():
        if nombre_empresa.upper() in nombre_mapeado.upper():
            return suc_id
        if nombre_mapeado.upper() in nombre_empresa.upper():
            return suc_id
    
    return None


def guardar_gastos_en_db(supabase, df_gastos, usuario=None):
    """
    Guarda los gastos procesados en la base de datos
    Mapea automáticamente nombres de empresas a sucursal_id
    IMPORTANTE: Usa la fecha de cada registro del CSV, no una fecha global
    
    Usa mapeo HÍBRIDO:
    1. Mapeo manual (tabla mapeo_sucursales_csv o hardcoded)
    2. Mapeo automático (tabla sucursales con estrategias)
    
    Retorna:
    --------
    dict con 'exitosos': int, 'errores': list, 'sin_sucursal': list, 'sin_fecha': list, 'duplicados': list
    """
    exitosos = 0
    errores = []
    sin_sucursal = []
    sin_fecha = []
    duplicados = []  # Registros que ya existen en DB
    
    # Obtener AMBOS mapeos
    mapeo_manual = obtener_mapeo_manual(supabase)
    mapeo_automatico = crear_mapeo_sucursales(supabase)
    
    try:
        for idx, row in df_gastos.iterrows():
            try:
                # Obtener nombre de la empresa
                nombre_empresa = row.get('Empresa', '')
                
                # Mapear a sucursal_id usando mapeo híbrido
                sucursal_id = obtener_sucursal_id_desde_nombre(
                    nombre_empresa, 
                    mapeo_manual, 
                    mapeo_automatico
                )
                
                if sucursal_id is None:
                    sin_sucursal.append({
                        'fila': idx + 1,
                        'empresa': nombre_empresa,
                        'fecha': row.get('Fecha', ''),
                        'total': row.get('TOTAL_GASTO', 0),
                        'sugerencia': '💡 Verifica que la sucursal exista en Supabase'
                    })
                    continue
                
                # IMPORTANTE: Extraer mes y año de la fecha del CSV
                # Intentar primero con 'Fecha C.' (Fecha de Contabilización)
                fecha_contable = None
                if 'Fecha C.' in row and pd.notna(row['Fecha C.']):
                    try:
                        fecha_contable = pd.to_datetime(row['Fecha C.'], format='%d/%m/%Y', errors='coerce')
                    except:
                        pass
                
                # Si no hay 'Fecha C.', usar 'Fecha E.' (Fecha de Emisión)
                if fecha_contable is None or pd.isna(fecha_contable):
                    if 'Fecha E.' in row and pd.notna(row['Fecha E.']):
                        try:
                            fecha_contable = pd.to_datetime(row['Fecha E.'], format='%d/%m/%Y', errors='coerce')
                        except:
                            pass
                
                # Si no hay ninguna fecha válida, usar la fecha procesada 'Fecha'
                if fecha_contable is None or pd.isna(fecha_contable):
                    if 'Fecha' in row and pd.notna(row['Fecha']):
                        fecha_contable = row['Fecha']
                
                # Verificar que tenemos una fecha válida
                if fecha_contable is None or pd.isna(fecha_contable):
                    sin_fecha.append({
                        'fila': idx + 1,
                        'empresa': nombre_empresa,
                        'fecha_e': row.get('Fecha E.', ''),
                        'fecha_c': row.get('Fecha C.', ''),
                        'total': row.get('TOTAL_GASTO', 0)
                    })
                    continue
                
                # Extraer mes y año de la fecha
                mes = fecha_contable.month
                anio = fecha_contable.year
                
                # Función auxiliar para convertir valores de forma segura
                def convertir_a_float_seguro(valor):
                    """Convierte un valor a float, reemplazando nan, inf, None con 0"""
                    try:
                        if pd.isna(valor) or valor is None:
                            return 0.0
                        val_float = float(valor)
                        # Verificar si es nan, inf o -inf
                        if not pd.isna(val_float) and val_float != float('inf') and val_float != float('-inf'):
                            return val_float
                        else:
                            return 0.0
                    except (ValueError, TypeError):
                        return 0.0
                
                # Preparar datos para inserción con conversión segura
                gasto_data = {
                    'sucursal_id': sucursal_id,
                    'mes': mes,  # Extraído de la fecha del CSV
                    'anio': anio,  # Extraído de la fecha del CSV
                    'fecha': str(fecha_contable.date()),
                    'tipo_comprobante': str(row.get('Tipo Comprobante', '')) if pd.notna(row.get('Tipo Comprobante')) else '',
                    'numero_comprobante': str(row.get('Comprobante', '')) if pd.notna(row.get('Comprobante')) else '',
                    'proveedor': str(row.get('Proveedor', '')) if pd.notna(row.get('Proveedor')) else '',
                    'cuit': str(row.get('Cuit', '')) if pd.notna(row.get('Cuit')) else '',
                    'rubro': str(row.get('Rubro', '')) if pd.notna(row.get('Rubro')) else '',
                    'subrubro': str(row.get('Subrubro', '')) if pd.notna(row.get('Subrubro')) else '',
                    'neto': convertir_a_float_seguro(row['NETO']),
                    'iva_percepciones': convertir_a_float_seguro(row['IVA_PERCEPCIONES']),
                    'total': convertir_a_float_seguro(row['TOTAL_GASTO']),
                    # Detalles de importes con conversión segura
                    'importe_neto_21': convertir_a_float_seguro(row.get('Importe Neto 21', 0)),
                    'importe_neto_10_5': convertir_a_float_seguro(row.get('Importe Neto 10_5', 0)),
                    'importe_neto_27': convertir_a_float_seguro(row.get('Importe Neto 27', 0)),
                    'impuestos_internos': convertir_a_float_seguro(row.get('Impuestos Internos', 0)),
                    'percepcion_iva': convertir_a_float_seguro(row.get('Percepcion Iva', 0)),
                    'otras_percepciones': convertir_a_float_seguro(row.get('Otras Percepciones', 0)),
                    'percepcion_iibb': convertir_a_float_seguro(row.get('Percepcion IIBB', 0)),
                    'iva_21': convertir_a_float_seguro(row.get('Iva 21', 0)),
                    'iva_10_5': convertir_a_float_seguro(row.get('Iva 10,5', 0)),
                    'iva_27': convertir_a_float_seguro(row.get('Iva 27', 0)),
                    'usuario_importacion': usuario
                }
                
                # Insertar en base de datos
                supabase.table("gastos_mensuales").insert(gasto_data).execute()
                exitosos += 1
                
            except Exception as e:
                error_str = str(e)
                # Detectar duplicados
                if 'duplicate key' in error_str.lower() and 'unique_gasto_registro' in error_str.lower():
                    duplicados.append({
                        'fila': idx + 1,
                        'sucursal_id': sucursal_id,
                        'empresa': nombre_empresa,
                        'proveedor': row.get('Proveedor', ''),
                        'total': row.get('TOTAL_GASTO', 0),
                        'fecha': str(fecha_contable.date()) if fecha_contable else ''
                    })
                else:
                    # Error real (no duplicado)
                    errores.append(f"Fila {idx + 1}: {error_str}")
        
        return {
            'exitosos': exitosos,
            'errores': errores,
            'sin_sucursal': sin_sucursal,
            'sin_fecha': sin_fecha,
            'duplicados': duplicados
        }
        
    except Exception as e:
        return {
            'exitosos': exitosos,
            'errores': [f"Error general: {str(e)}"] + errores,
            'sin_sucursal': sin_sucursal,
            'sin_fecha': sin_fecha
        }


def eliminar_gastos_periodo(supabase, sucursal_id, mes, anio):
    """
    Elimina todos los gastos de un período específico
    """
    try:
        supabase.table("gastos_mensuales")\
            .delete()\
            .eq("sucursal_id", sucursal_id)\
            .eq("mes", mes)\
            .eq("anio", anio)\
            .execute()
        return True
    except Exception as e:
        st.error(f"❌ Error eliminando gastos: {str(e)}")
        return False


@st.cache_data(ttl=30)  # 🚀 OPTIMIZACIÓN: Cachear por 30 segundos
def obtener_gastos_db(_supabase, mes, anio, sucursal_id=None):
    """
    Obtiene los gastos desde la base de datos
    
    OPTIMIZADO: Resultados se cachean por 30 segundos para mejorar rendimiento.
    El guión bajo (_supabase) indica a Streamlit que no use este parámetro para el caché.
    """
    try:
        query = _supabase.table("gastos_mensuales").select("*")
        
        query = query.eq("mes", mes)
        query = query.eq("anio", anio)
        
        if sucursal_id is not None:
            query = query.eq("sucursal_id", sucursal_id)
        
        result = query.execute()
        
        if result.data:
            return pd.DataFrame(result.data)
        else:
            return pd.DataFrame()
            
    except Exception as e:
        st.error(f"❌ Error obteniendo gastos de DB: {str(e)}")
        return pd.DataFrame()


@st.cache_data(ttl=30)  # 🚀 OPTIMIZACIÓN: Cachear por 30 segundos
def obtener_ingresos_mensuales(_supabase, mes, anio, sucursal_id=None):
    """
    Obtiene los ingresos mensuales de la base de datos de cajas_diarias
    
    OPTIMIZADO: Resultados se cachean por 30 segundos para mejorar rendimiento.
    El guión bajo (_supabase) indica a Streamlit que no use este parámetro para el caché.
    """
    try:
        # Construir fechas de inicio y fin del mes
        primer_dia = date(anio, mes, 1)
        ultimo_dia = date(anio, mes, calendar.monthrange(anio, mes)[1])
        
        # Query base
        query = _supabase.table("movimientos_diarios").select("*")
        
        # Filtrar por fechas
        query = query.gte("fecha", str(primer_dia))
        query = query.lte("fecha", str(ultimo_dia))
        
        # Filtrar por sucursal si se especifica
        if sucursal_id is not None:
            query = query.eq("sucursal_id", sucursal_id)
        
        # Filtrar solo ventas (ingresos)
        query = query.eq("tipo", "venta")  # ✅ CORREGIDO: era "ingreso"
        
        # Ejecutar query
        result = query.execute()
        
        if result.data:
            df = pd.DataFrame(result.data)
            return df
        else:
            return pd.DataFrame()
            
    except Exception as e:
        st.error(f"❌ Error obteniendo ingresos: {str(e)}")
        return pd.DataFrame()


def generar_excel_con_detalle(df_ingresos, df_gastos, total_ingresos, total_gastos, resultado, margen_porcentaje, sucursal_nombre, mes_seleccionado, anio_seleccionado):
    """
    Genera un archivo Excel con múltiples pestañas:
    - Resumen: Estado de resultados resumido
    - Detalle Movimientos: Todos los movimientos que componen el estado de resultados
    """
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Crear el archivo Excel en memoria
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # --- PESTANA 1: RESUMEN ---
        meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        nombre_mes = meses[mes_seleccionado]
        
        # Crear DataFrame de resumen
        resumen_data = []
        
        # Encabezado
        resumen_data.append(['ESTADO DE RESULTADOS', ''])
        resumen_data.append([f'{nombre_mes.upper()} {anio_seleccionado}', ''])
        resumen_data.append([sucursal_nombre.upper(), ''])
        resumen_data.append(['', ''])
        
        # INGRESOS
        resumen_data.append(['VENTAS/INGRESOS', ''])
        if not df_ingresos.empty:
            if 'categoria' in df_ingresos.columns:
                ingresos_por_categoria = df_ingresos.groupby('categoria')['monto'].sum()
                for categoria, monto in ingresos_por_categoria.items():
                    resumen_data.append([f'  {categoria.title()}', monto])
            else:
                resumen_data.append(['  Ventas Salon', total_ingresos])
        resumen_data.append(['TOTAL INGRESOS', total_ingresos])
        resumen_data.append(['', ''])
        
        # GASTOS
        resumen_data.append(['COMPRAS/EGRESOS', ''])
        if 'rubro' in df_gastos.columns:
            gastos_agrupados = df_gastos.groupby('rubro')['total'].sum().sort_values(ascending=False)
            
            # CMC
            alimentos = gastos_agrupados[gastos_agrupados.index.str.contains('ALIMENTOS', case=False, na=False)].sum()
            bebidas = gastos_agrupados[gastos_agrupados.index.str.contains('BEBIDAS', case=False, na=False)].sum()
            total_cmc = alimentos + bebidas
            
            if total_cmc > 0:
                resumen_data.append(['Costo de Mercaderia Consumida (CMC)', total_cmc])
                if alimentos > 0:
                    resumen_data.append(['  Alimentos', alimentos])
                if bebidas > 0:
                    resumen_data.append(['  Bebidas', bebidas])
            
            # Personal
            sueldos = gastos_agrupados[gastos_agrupados.index.str.contains('SUELDOS', case=False, na=False)].sum()
            cargas_sociales = gastos_agrupados[gastos_agrupados.index.str.contains('CARGAS SOCIALES', case=False, na=False)].sum()
            total_personal = sueldos + cargas_sociales
            
            if total_personal > 0:
                resumen_data.append(['Gastos de Personal', total_personal])
                if sueldos > 0:
                    resumen_data.append(['  Sueldos y Salarios', sueldos])
                if cargas_sociales > 0:
                    resumen_data.append(['  Cargas Sociales', cargas_sociales])
            
            # Operativos
            alquiler = gastos_agrupados[gastos_agrupados.index.str.contains('ALQUILER', case=False, na=False)].sum()
            servicios = gastos_agrupados[gastos_agrupados.index.str.contains('SERVICIOS', case=False, na=False)].sum()
            mantenimiento = gastos_agrupados[gastos_agrupados.index.str.contains('MANTENIMIENTO', case=False, na=False)].sum()
            publicidad = gastos_agrupados[gastos_agrupados.index.str.contains('PUBLICIDAD', case=False, na=False)].sum()
            insumos = gastos_agrupados[gastos_agrupados.index.str.contains('INSUMOS', case=False, na=False)].sum()
            materiales = gastos_agrupados[gastos_agrupados.index.str.contains('MATERIALES', case=False, na=False)].sum()
            
            # Obtener otras categorias
            categorias_principales_keywords = ['ALIMENTOS', 'BEBIDAS', 'SUELDOS', 'CARGAS SOCIALES', 
                                               'ALQUILER', 'SERVICIOS', 'MANTENIMIENTO', 'PUBLICIDAD', 
                                               'INSUMOS', 'MATERIALES', 'IMPUESTOS']
            
            otras_categorias = []
            for categoria, monto in gastos_agrupados.items():
                es_principal = any(keyword in categoria.upper() for keyword in categorias_principales_keywords)
                if not es_principal:
                    otras_categorias.append((categoria, monto))
            
            total_otros_gastos = sum(monto for _, monto in otras_categorias)
            
            total_operativos = alquiler + servicios + mantenimiento + publicidad + total_otros_gastos + insumos + materiales
            
            if total_operativos > 0:
                resumen_data.append(['Gastos Operativos', total_operativos])
                if alquiler > 0:
                    resumen_data.append(['  Alquileres', alquiler])
                if servicios > 0:
                    resumen_data.append(['  Servicios', servicios])
                if mantenimiento > 0:
                    resumen_data.append(['  Mantenimiento', mantenimiento])
                if publicidad > 0:
                    resumen_data.append(['  Publicidad', publicidad])
                if total_otros_gastos > 0:
                    resumen_data.append(['  Otros Gastos', total_otros_gastos])
                    for categoria, monto in otras_categorias:
                        resumen_data.append([f'    {categoria}', monto])
                if insumos > 0:
                    resumen_data.append(['  Insumos y Descartables', insumos])
                if materiales > 0:
                    resumen_data.append(['  Materiales', materiales])
            
            # Impuestos
            impuestos = gastos_agrupados[gastos_agrupados.index.str.contains('IMPUESTOS', case=False, na=False)].sum()
            if impuestos > 0:
                resumen_data.append(['Impuestos y Tasas', impuestos])
        
        resumen_data.append(['TOTAL EGRESOS', total_gastos])
        resumen_data.append(['', ''])
        
        # RESULTADO
        resumen_data.append(['RESULTADO OPERATIVO', resultado])
        resumen_data.append(['MARGEN %', f'{margen_porcentaje:.2f}%'])
        
        df_resumen = pd.DataFrame(resumen_data, columns=['Concepto', 'Monto'])
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
        
        # Formatear pestana de resumen
        worksheet_resumen = writer.sheets['Resumen']
        
        # Ajustar anchos de columna
        worksheet_resumen.column_dimensions['A'].width = 40
        worksheet_resumen.column_dimensions['B'].width = 20
        
        # Aplicar estilos
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for row in worksheet_resumen.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
        
        # Formato de moneda para columna B (excepto encabezados y porcentajes)
        for row_idx, row in enumerate(worksheet_resumen.iter_rows(min_row=2), start=2):
            cell_concepto = row[0]
            cell_monto = row[1]
            
            # Negrita para totales
            if cell_concepto.value and any(keyword in str(cell_concepto.value).upper() 
                                           for keyword in ['TOTAL', 'RESULTADO', 'VENTAS/INGRESOS', 'COMPRAS/EGRESOS']):
                cell_concepto.font = Font(bold=True, size=11)
                cell_monto.font = Font(bold=True, size=11)
            
            # Formato de moneda
            if cell_monto.value and isinstance(cell_monto.value, (int, float)) and 'MARGEN' not in str(cell_concepto.value).upper():
                cell_monto.number_format = '$#,##0.00'
            
            # Alineacion
            cell_concepto.alignment = Alignment(horizontal='left')
            cell_monto.alignment = Alignment(horizontal='right')
        
        # --- PESTANA 2: DETALLE DE MOVIMIENTOS ---
        # Preparar datos de gastos para el detalle
        if not df_gastos.empty:
            df_detalle_gastos = df_gastos.copy()
            
            # Seleccionar y renombrar columnas relevantes
            columnas_detalle = []
            rename_dict = {}
            
            # Mapeo de columnas estandar
            if 'fecha' in df_detalle_gastos.columns:
                columnas_detalle.append('fecha')
                rename_dict['fecha'] = 'Fecha'
            
            if 'empresa' in df_detalle_gastos.columns:
                columnas_detalle.append('empresa')
                rename_dict['empresa'] = 'Empresa'
            elif 'proveedor' in df_detalle_gastos.columns:
                columnas_detalle.append('proveedor')
                rename_dict['proveedor'] = 'Proveedor'
            
            if 'tipo_comprobante' in df_detalle_gastos.columns:
                columnas_detalle.append('tipo_comprobante')
                rename_dict['tipo_comprobante'] = 'Tipo Comprobante'
            
            if 'numero_factura' in df_detalle_gastos.columns:
                columnas_detalle.append('numero_factura')
                rename_dict['numero_factura'] = 'Nro Factura'
            
            if 'rubro' in df_detalle_gastos.columns:
                columnas_detalle.append('rubro')
                rename_dict['rubro'] = 'Categoria'
            elif 'categoria' in df_detalle_gastos.columns:
                columnas_detalle.append('categoria')
                rename_dict['categoria'] = 'Categoria'
            
            if 'neto' in df_detalle_gastos.columns:
                columnas_detalle.append('neto')
                rename_dict['neto'] = 'Neto'
            
            if 'iva_percepciones' in df_detalle_gastos.columns:
                columnas_detalle.append('iva_percepciones')
                rename_dict['iva_percepciones'] = 'IVA y Percepciones'
            
            if 'total' in df_detalle_gastos.columns:
                columnas_detalle.append('total')
                rename_dict['total'] = 'Total'
            
            # Filtrar y renombrar
            df_detalle_export = df_detalle_gastos[columnas_detalle].copy()
            df_detalle_export = df_detalle_export.rename(columns=rename_dict)
            
            # Ordenar por fecha
            if 'Fecha' in df_detalle_export.columns:
                df_detalle_export = df_detalle_export.sort_values('Fecha')
            
            # Escribir a Excel
            df_detalle_export.to_excel(writer, sheet_name='Detalle Movimientos', index=False)
            
            # Formatear pestana de detalle
            worksheet_detalle = writer.sheets['Detalle Movimientos']
            
            # Ajustar anchos de columna
            for column in worksheet_detalle.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet_detalle.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar estilos al encabezado
            for cell in worksheet_detalle[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Formato de moneda para columnas de monto
            for row in worksheet_detalle.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.column > 5:
                        cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal='left')
            
            # Agregar fila de totales al final
            ultima_fila = worksheet_detalle.max_row + 1
            worksheet_detalle.cell(row=ultima_fila, column=1).value = 'TOTAL GENERAL'
            worksheet_detalle.cell(row=ultima_fila, column=1).font = Font(bold=True, size=11)
            
            # Buscar columna de total
            col_total = None
            for idx, cell in enumerate(worksheet_detalle[1], start=1):
                if cell.value == 'Total':
                    col_total = idx
                    break
            
            if col_total:
                total_formula = f'=SUM({openpyxl.utils.get_column_letter(col_total)}2:{openpyxl.utils.get_column_letter(col_total)}{ultima_fila-1})'
                cell_total = worksheet_detalle.cell(row=ultima_fila, column=col_total)
                cell_total.value = total_formula
                cell_total.font = Font(bold=True, size=11)
                cell_total.number_format = '$#,##0.00'
    
    output.seek(0)
    return output


@st.cache_data(ttl=30)  # OPTIMIZACION: Cachear por 30 segundos
def obtener_evolucion_historica(_supabase, sucursal_id, meses_atras=12):
    """
    Obtiene la evolución histórica de gastos e ingresos
    
    OPTIMIZADO: Resultados se cachean por 30 segundos para mejorar rendimiento.
    El guión bajo (_supabase) indica a Streamlit que no use este parámetro para el caché.
    """
    try:
        # Obtener gastos históricos
        result_gastos = _supabase.table("gastos_mensuales")\
            .select("anio, mes, total")\
            .eq("sucursal_id", sucursal_id)\
            .execute()
        
        if not result_gastos.data:
            return pd.DataFrame()
        
        # Convertir a DataFrame y agrupar
        df_gastos = pd.DataFrame(result_gastos.data)
        df_gastos_agg = df_gastos.groupby(['anio', 'mes'])['total'].sum().reset_index()
        df_gastos_agg.columns = ['anio', 'mes', 'total_gastos']
        
        # Crear columna de período
        df_gastos_agg['periodo'] = pd.to_datetime(
            df_gastos_agg['anio'].astype(str) + '-' + 
            df_gastos_agg['mes'].astype(str).str.zfill(2) + '-01'
        )
        
        # Obtener ingresos históricos
        fecha_limite = pd.Timestamp.now() - pd.DateOffset(months=meses_atras)
        
        result_ingresos = _supabase.table("movimientos_diarios")\
            .select("fecha, monto")\
            .eq("sucursal_id", sucursal_id)\
            .eq("tipo", "venta")\
            .gte("fecha", str(fecha_limite.date()))\
            .execute()  # ✅ CORREGIDO: tipo era "ingreso", ahora "venta"
        
        if result_ingresos.data:
            df_ingresos = pd.DataFrame(result_ingresos.data)
            df_ingresos['fecha'] = pd.to_datetime(df_ingresos['fecha'])
            df_ingresos['anio'] = df_ingresos['fecha'].dt.year
            df_ingresos['mes'] = df_ingresos['fecha'].dt.month
            df_ingresos['periodo'] = df_ingresos['fecha'].dt.to_period('M').dt.to_timestamp()
            
            df_ingresos_agg = df_ingresos.groupby(['anio', 'mes', 'periodo'])['monto'].sum().reset_index()
            df_ingresos_agg.columns = ['anio', 'mes', 'periodo', 'total_ingresos']
        else:
            df_ingresos_agg = pd.DataFrame(columns=['anio', 'mes', 'periodo', 'total_ingresos'])
        
        # Merge de gastos e ingresos
        df_evolucion = pd.merge(
            df_gastos_agg,
            df_ingresos_agg,
            on=['anio', 'mes', 'periodo'],
            how='outer'
        ).fillna(0)
        
        # Calcular resultado y margen
        df_evolucion['resultado'] = df_evolucion['total_ingresos'] - df_evolucion['total_gastos']
        df_evolucion['margen'] = (df_evolucion['resultado'] / df_evolucion['total_ingresos'] * 100).fillna(0)
        
        # Ordenar por período
        df_evolucion = df_evolucion.sort_values('periodo', ascending=False)
        
        # Limitar a últimos N meses
        df_evolucion = df_evolucion.head(meses_atras)
        
        return df_evolucion.sort_values('periodo')
        
    except Exception as e:
        st.error(f"❌ Error obteniendo evolución histórica: {str(e)}")
        return pd.DataFrame()


def limpiar_cache_pl_simples():
    """
    Limpia el caché de las consultas de P&L Simples.
    
    🚀 OPTIMIZACIÓN: Llamar después de importar CSV o cuando se necesiten datos frescos.
    """
    try:
        obtener_gastos_db.clear()
        obtener_ingresos_mensuales.clear()
        obtener_evolucion_historica.clear()
        return True
    except Exception as e:
        st.warning(f"⚠️ No se pudo limpiar el caché: {str(e)}")
        return False


def calcular_benchmarks_gastronomia():
    """
    Retorna benchmarks estándar del rubro gastronómico
    Basado en estadísticas de la industria argentina
    """
    return {
        'ALIMENTOS': {'porcentaje_ideal': 30.0, 'rango_min': 25.0, 'rango_max': 35.0},
        'BEBIDAS': {'porcentaje_ideal': 8.0, 'rango_min': 6.0, 'rango_max': 10.0},
        'SUELDOS': {'porcentaje_ideal': 30.0, 'rango_min': 25.0, 'rango_max': 35.0},
        'SERVICIOS': {'porcentaje_ideal': 15.0, 'rango_min': 12.0, 'rango_max': 18.0},
        'ALQUILER': {'porcentaje_ideal': 8.0, 'rango_min': 6.0, 'rango_max': 10.0},
        'INSUMOS Y DESCARTABLES': {'porcentaje_ideal': 4.0, 'rango_min': 3.0, 'rango_max': 5.0},
        'IMPUESTOS': {'porcentaje_ideal': 5.0, 'rango_min': 4.0, 'rango_max': 6.0},
        'MATERIALES': {'porcentaje_ideal': 3.0, 'rango_min': 2.0, 'rango_max': 5.0}
    }


def analizar_composicion_gastos(df_gastos, ingresos_totales):
    """
    Analiza la composición de gastos y compara con benchmarks
    """
    # Determinar columna de totales
    col_total = 'TOTAL_GASTO' if 'TOTAL_GASTO' in df_gastos.columns else 'total'
    col_rubro = 'Rubro' if 'Rubro' in df_gastos.columns else 'rubro'
    
    # Agrupar por rubro
    gastos_por_rubro = df_gastos.groupby(col_rubro)[col_total].sum().sort_values(ascending=False)
    
    # Calcular porcentajes sobre ingresos
    total_gastos = gastos_por_rubro.sum()
    
    analisis = {
        'total_gastos': total_gastos,
        'ingresos': ingresos_totales,
        'resultado': ingresos_totales - total_gastos,
        'margen_porcentaje': ((ingresos_totales - total_gastos) / ingresos_totales * 100) if ingresos_totales > 0 else 0,
        'rubros': []
    }
    
    benchmarks = calcular_benchmarks_gastronomia()
    
    for rubro, gasto in gastos_por_rubro.items():
        porcentaje_real = (gasto / ingresos_totales * 100) if ingresos_totales > 0 else 0
        
        # Buscar benchmark
        rubro_upper = str(rubro).upper()
        benchmark = None
        estado = "neutral"
        
        # Buscar benchmark exacto o parcial
        for key in benchmarks.keys():
            if key in rubro_upper or rubro_upper in key:
                benchmark = benchmarks[key]
                
                # Evaluar estado
                if porcentaje_real < benchmark['rango_min']:
                    estado = "bajo"
                elif porcentaje_real > benchmark['rango_max']:
                    estado = "alto"
                else:
                    estado = "ok"
                break
        
        analisis['rubros'].append({
            'rubro': rubro,
            'gasto': gasto,
            'porcentaje_real': porcentaje_real,
            'benchmark': benchmark,
            'estado': estado
        })
    
    return analisis


# ==================== INTERFAZ STREAMLIT ====================

def mostrar_tab_importacion(supabase, sucursales, mes_seleccionado, anio_seleccionado, sucursal_seleccionada):
    """
    Tab de importación de gastos desde CSV
    Soporta múltiples sucursales en un mismo CSV con mapeo automático
    Las fechas se extraen automáticamente de cada registro del CSV
    """
    st.subheader("📁 Importar Gastos desde CSV")
    
    st.info("💡 **Importación Inteligente**: El sistema detecta automáticamente las sucursales Y las fechas de cada gasto desde el CSV. No necesitas seleccionar mes/año manualmente.")
    
    # Crear AMBOS mapeos
    mapeo_manual = obtener_mapeo_manual(supabase)
    mapeo_automatico = crear_mapeo_sucursales(supabase)
    
    if not mapeo_automatico:
        st.error("❌ No se pudieron cargar las sucursales. Verifica la conexión a Supabase.")
        return
    
    # Mostrar información de mapeo
    with st.expander("🏪 Información de Mapeo"):
        if mapeo_manual:
            st.write("✅ **Mapeo Manual Activo** (prioridad alta)")
            st.write(f"   - {len(set(mapeo_manual.values()))} empresas CSV -> sucursales")
        else:
            st.write("⚠️ **Mapeo Manual No Configurado** (solo mapeo automático)")
        
        st.write("")
        st.write("📋 **Sucursales disponibles:**")
        for nombre, suc_id in mapeo_automatico.items():
            if ' ' in nombre or '.' in nombre:  # Mostrar solo nombres completos
                st.write(f"- {nombre} (ID: {suc_id})")
    
    # Formulario de importación
    archivo_gastos = st.file_uploader(
        "Selecciona el archivo CSV de gastos del mes",
        type=['csv'],
        help="Archivo CSV exportado del sistema de facturas de compra"
    )
    
    if archivo_gastos is not None:
        # Procesar archivo
        with st.spinner("Procesando archivo CSV..."):
            df_gastos = procesar_archivo_gastos(archivo_gastos)
        
        if df_gastos is not None and len(df_gastos) > 0:
            st.success(f"✅ Archivo procesado: {len(df_gastos)} registros detectados")
            
            # Analizar períodos en el CSV
            if 'Fecha' in df_gastos.columns and df_gastos['Fecha'].notna().any():
                fechas_validas = df_gastos[df_gastos['Fecha'].notna()]['Fecha']
                fecha_min = fechas_validas.min()
                fecha_max = fechas_validas.max()
                
                periodos = df_gastos[df_gastos['Fecha'].notna()].copy()
                periodos['periodo'] = periodos['Fecha'].dt.to_period('M')
                periodos_unicos = periodos['periodo'].unique()
                
                st.info(f"📅 **Períodos detectados en el CSV**: {fecha_min.strftime('%d/%m/%Y')} hasta {fecha_max.strftime('%d/%m/%Y')} ({len(periodos_unicos)} mes(es))")
                
                # Mostrar resumen por período
                with st.expander("📊 Ver distribución por período"):
                    resumen_periodo = periodos.groupby('periodo').size().reset_index(name='registros')
                    resumen_periodo['periodo_str'] = resumen_periodo['periodo'].astype(str)
                    st.dataframe(
                        resumen_periodo[['periodo_str', 'registros']].rename(columns={
                            'periodo_str': 'Período',
                            'registros': 'Cantidad de Registros'
                        }),
                        hide_index=True
                    )
            
            # Analizar sucursales en el CSV
            st.markdown("---")
            st.subheader("🔍 Análisis de Sucursales Detectadas")
            
            empresas_detectadas = df_gastos['Empresa'].value_counts()
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.metric("Total de Empresas en CSV", len(empresas_detectadas))
            
            # Verificar mapeo para cada empresa
            empresas_mapeadas = []
            empresas_sin_mapear = []
            
            for empresa, cantidad in empresas_detectadas.items():
                sucursal_id = obtener_sucursal_id_desde_nombre(empresa, mapeo_manual, mapeo_automatico)
                total_empresa = df_gastos[df_gastos['Empresa'] == empresa]['TOTAL_GASTO'].sum()
                
                if sucursal_id:
                    # Encontrar nombre de sucursal
                    nombre_sucursal = next((k for k, v in mapeo_automatico.items() if v == sucursal_id and (' ' in k or '.' in k)), empresa)
                    empresas_mapeadas.append({
                        'CSV': empresa,
                        'Sucursal': nombre_sucursal,
                        'Registros': cantidad,
                        'Total': total_empresa
                    })
                else:
                    empresas_sin_mapear.append({
                        'Empresa': empresa,
                        'Registros': cantidad,
                        'Total': total_empresa
                    })
            
            with col2:
                st.metric("Empresas Mapeadas ✅", len(empresas_mapeadas))
            
            # Mostrar empresas mapeadas
            if empresas_mapeadas:
                st.success("✅ **Empresas que se importarán correctamente:**")
                df_mapeadas = pd.DataFrame(empresas_mapeadas)
                df_mapeadas['Total Formateado'] = df_mapeadas['Total'].apply(
                    lambda x: f"${x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                )
                st.dataframe(
                    df_mapeadas[['CSV', 'Sucursal', 'Registros', 'Total Formateado']],
                    hide_index=True,
                    width="stretch"
                )
            
            # Mostrar empresas sin mapear
            if empresas_sin_mapear:
                st.error("⚠️ **Empresas SIN MAPEAR (no se importarán):**")
                df_sin_mapear = pd.DataFrame(empresas_sin_mapear)
                df_sin_mapear['Total Formateado'] = df_sin_mapear['Total'].apply(
                    lambda x: f"${x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                )
                st.dataframe(
                    df_sin_mapear[['Empresa', 'Registros', 'Total Formateado']],
                    hide_index=True,
                    width="stretch"
                )
                
                st.warning("💡 **Solución**: Crea estas sucursales en el sistema o ajusta los nombres en el CSV para que coincidan.")
            
            # Vista previa de datos
            with st.expander("👁️ Vista previa de primeros 10 registros"):
                st.dataframe(
                    df_gastos[['Empresa', 'Fecha', 'Rubro', 'Subrubro', 'Proveedor', 
                              'NETO', 'IVA_PERCEPCIONES', 'TOTAL_GASTO']].head(10),
                    hide_index=True
                )
            
            # Verificar si ya existen gastos para alguna sucursal en los períodos del CSV
            st.markdown("---")
            st.subheader("⚠️ Verificación de Duplicados")
            
            # Obtener períodos únicos del CSV
            periodos_csv = set()
            if 'Fecha' in df_gastos.columns:
                for fecha in df_gastos[df_gastos['Fecha'].notna()]['Fecha']:
                    periodos_csv.add((fecha.year, fecha.month))
            
            gastos_existentes_info = []
            for empresa_info in empresas_mapeadas:
                sucursal_id = obtener_sucursal_id_desde_nombre(empresa_info['CSV'], mapeo_manual, mapeo_automatico)
                if sucursal_id:
                    for anio, mes in periodos_csv:
                        gastos_existentes = verificar_gastos_existentes(supabase, sucursal_id, mes, anio)
                        if gastos_existentes['existe']:
                            gastos_existentes_info.append({
                                'sucursal': empresa_info['Sucursal'],
                                'periodo': f"{mes:02d}/{anio}",
                                'cantidad': gastos_existentes['cantidad'],
                                'total': gastos_existentes['total'],
                                'sucursal_id': sucursal_id,
                                'mes': mes,
                                'anio': anio
                            })
            
            if gastos_existentes_info:
                st.warning("⚠️ **Ya existen gastos para algunos períodos:**")
                for info in gastos_existentes_info:
                    st.write(f"- **{info['sucursal']}** ({info['periodo']}): {info['cantidad']} registros (${info['total']:,.2f})".replace(',', 'X').replace('.', ',').replace('X', '.'))
                
                st.write("")
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("🔄 Reemplazar TODOS los gastos existentes", type="secondary"):
                        with st.spinner("Eliminando gastos existentes..."):
                            for info in gastos_existentes_info:
                                eliminar_gastos_periodo(supabase, info['sucursal_id'], info['mes'], info['anio'])
                            st.success("✅ Gastos eliminados. Puedes importar nuevos datos.")
                            st.session_state['gastos_eliminados'] = True
                            st.rerun()
                
                with col2:
                    if st.button("❌ Cancelar y mantener existentes"):
                        st.info("Operación cancelada. Los gastos existentes se mantienen.")
                        return
                
                if 'gastos_eliminados' not in st.session_state or not st.session_state['gastos_eliminados']:
                    st.info("⚠️ Debes decidir si reemplazar o cancelar antes de continuar.")
                    st.stop()
            
            # Resumen antes de guardar
            st.markdown("---")
            st.subheader("📊 Resumen de la Importación")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Registros", len(df_gastos))
            with col2:
                registros_a_importar = sum(e['Registros'] for e in empresas_mapeadas)
                st.metric("Se Importarán", registros_a_importar)
            with col3:
                st.metric("Total Neto", f"${df_gastos['NETO'].sum():,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
            with col4:
                st.metric("Total General", f"${df_gastos['TOTAL_GASTO'].sum():,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
            
            # Botón de importación
            st.markdown("---")
            usuario_actual = st.session_state.get('usuario', {}).get('usuario', 'desconocido')
            
            puede_importar = len(empresas_mapeadas) > 0
            
            if not puede_importar:
                st.error("❌ No se puede importar porque ninguna empresa fue mapeada correctamente.")
                st.info("💡 Crea las sucursales en el sistema o ajusta los nombres en el CSV.")
            else:
                if st.button("💾 Guardar en Base de Datos", type="primary", width="stretch", disabled=not puede_importar):
                    with st.spinner("Guardando gastos en la base de datos..."):
                        resultado = guardar_gastos_en_db(
                            supabase, 
                            df_gastos,
                            usuario_actual
                        )
                        
                        if resultado['exitosos'] > 0:
                            st.success(f"✅ {resultado['exitosos']} registros guardados exitosamente")
                            
                            # Mostrar detalle por sucursal
                            if empresas_mapeadas:
                                st.info("📊 **Registros guardados por sucursal:**")
                                for empresa_info in empresas_mapeadas:
                                    st.write(f"- {empresa_info['Sucursal']}: {empresa_info['Registros']} registros")
                            
                            # Limpiar cache y session_state
                            if 'gastos_eliminados' in st.session_state:
                                del st.session_state['gastos_eliminados']
                            
                            # 🚀 OPTIMIZACIÓN: Limpiar solo caché de P&L (más eficiente que limpiar todo)
                            limpiar_cache_pl_simples()
                            
                            st.info("💡 Los gastos fueron guardados con sus fechas originales del CSV. Ve a 'Análisis del Período' o 'Evolución Histórica'.")
                        
                        # Mostrar duplicados agrupados por sucursal
                        if resultado.get('duplicados'):
                            # Agrupar duplicados por sucursal
                            duplicados_por_sucursal = {}
                            for dup in resultado['duplicados']:
                                suc_id = dup['sucursal_id']
                                if suc_id not in duplicados_por_sucursal:
                                    # Buscar nombre de sucursal
                                    nombre_suc = next(
                                        (k for k, v in mapeo_automatico.items() if v == suc_id and (' ' in k or '.' in k)), 
                                        dup['empresa']
                                    )
                                    duplicados_por_sucursal[suc_id] = {
                                        'nombre': nombre_suc,
                                        'cantidad': 0
                                    }
                                duplicados_por_sucursal[suc_id]['cantidad'] += 1
                            
                            # Mostrar mensaje agrupado
                            st.warning(f"⚠️ **{len(resultado['duplicados'])} registros no se importaron porque ya existen en la BD:**")
                            for suc_id, info in duplicados_por_sucursal.items():
                                st.write(f"  • **{info['nombre']}**: {info['cantidad']} registro(s)")
                            
                            with st.expander("🔍 Ver detalle de duplicados"):
                                for dup in resultado['duplicados'][:20]:
                                    st.write(f"  - Fila {dup['fila']}: {dup['proveedor']} (${dup['total']:,.2f}) - {dup['fecha']}")
                                if len(resultado['duplicados']) > 20:
                                    st.write(f"  ... y {len(resultado['duplicados'])-20} más")
                        
                        if resultado.get('sin_fecha'):
                            st.warning(f"⚠️ {len(resultado['sin_fecha'])} registros sin fecha válida (no importados):")
                            for item in resultado['sin_fecha'][:10]:
                                st.write(f"  • Fila {item['fila']}: {item['empresa']} - Fecha E: {item.get('fecha_e')} - Fecha C: {item.get('fecha_c')}")
                            if len(resultado['sin_fecha']) > 10:
                                st.write(f"  ... y {len(resultado['sin_fecha'])-10} más")
                        
                        if resultado['sin_sucursal']:
                            st.warning(f"⚠️ {len(resultado['sin_sucursal'])} registros sin sucursal mapeada (no importados):")
                            for item in resultado['sin_sucursal'][:10]:
                                st.write(f"  • Fila {item['fila']}: {item['empresa']} (${item['total']:,.2f})".replace(',', 'X').replace('.', ',').replace('X', '.'))
                            if len(resultado['sin_sucursal']) > 10:
                                st.write(f"  ... y {len(resultado['sin_sucursal'])-10} más")
                        
                        if resultado['errores']:
                            st.error(f"❌ {len(resultado['errores'])} errores durante la importación:")
                            for error in resultado['errores'][:5]:
                                st.error(f"  • {error}")
                            if len(resultado['errores']) > 5:
                                st.error(f"  ... y {len(resultado['errores'])-5} errores más")
        
        else:
            st.error("❌ No se pudo procesar el archivo de gastos")


def mostrar_tab_analisis(supabase, sucursales, mes_seleccionado, anio_seleccionado, sucursal_seleccionada):
    """
    Tab de análisis del período actual con diseño profesional de informe financiero
    """
    # Header con botón de refrescar
    col_header, col_refresh = st.columns([4, 1])
    
    with col_header:
        st.subheader("📊 Análisis del Período")
    
    with col_refresh:
        if st.button("🔄 Refrescar", key="refresh_analisis", width="stretch", help="Actualizar datos desde la base de datos"):
            limpiar_cache_pl_simples()
            st.success("✅ Datos actualizados")
            st.rerun()
    
    sucursal_id = sucursal_seleccionada['id'] if sucursal_seleccionada else None
    sucursal_nombre = sucursal_seleccionada['nombre'] if sucursal_seleccionada else "Todas las sucursales"
    
    # Obtener datos de la DB
    with st.spinner("Cargando datos..."):
        df_gastos = obtener_gastos_db(supabase, mes_seleccionado, anio_seleccionado, sucursal_id)
        df_ingresos = obtener_ingresos_mensuales(supabase, mes_seleccionado, anio_seleccionado, sucursal_id)
    
    if df_gastos.empty:
        st.warning(f"⚠️ No hay gastos registrados para **{sucursal_nombre}** en **{mes_seleccionado}/{anio_seleccionado}**.")
        st.info("💡 Ve a la pestaña 'Importar Gastos' para cargar datos.")
        return
    
    # Verificar si los gastos son de sucursales diferentes a la seleccionada
    if sucursal_seleccionada:
        sucursales_en_gastos = df_gastos['sucursal_id'].unique()
        if sucursal_id not in sucursales_en_gastos:
            st.warning(f"⚠️ Los gastos de este período pertenecen a otras sucursales.")
            st.info(f"💡 Sucursales con gastos en {mes_seleccionado}/{anio_seleccionado}:")
            for suc_id in sucursales_en_gastos:
                suc = next((s for s in sucursales if s['id'] == suc_id), None)
                if suc:
                    cantidad = len(df_gastos[df_gastos['sucursal_id'] == suc_id])
                    st.write(f"   - {suc['nombre']}: {cantidad} registros")
            st.info("💡 Selecciona una de estas sucursales en el selector de arriba para ver su análisis.")
            return
    
    # Calcular totales
    total_ingresos = df_ingresos['monto'].sum() if not df_ingresos.empty else 0
    total_gastos = df_gastos['total'].sum()
    resultado = total_ingresos - total_gastos
    margen_porcentaje = (resultado / total_ingresos * 100) if total_ingresos > 0 else 0
    
    # INICIO: DISEÑO PROFESIONAL DE INFORME FINANCIERO
    st.markdown("---")
    
    # Título del informe
    meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    nombre_mes = meses[mes_seleccionado]
    
    st.markdown(f"""
    <div style="text-align: center; padding: 20px; background-color: #f8f9fa; border-radius: 10px; margin-bottom: 30px;">
        <h1 style="color: #2c3e50; margin-bottom: 10px;">ESTADO DE RESULTADOS</h1>
        <h2 style="color: #34495e; margin-bottom: 5px;">{nombre_mes.upper()} {anio_seleccionado}</h2>
        <h3 style="color: #7f8c8d; font-weight: normal;">{sucursal_nombre.upper()}</h3>
    </div>
    """, unsafe_allow_html=True)
    
    # KPIs principales en tarjetas profesionales
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        color_ingresos = "#27ae60" if total_ingresos > 0 else "#95a5a6"
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); 
                    padding: 20px; border-radius: 10px; text-align: center; 
                    border-left: 5px solid {color_ingresos}; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
            <div style="color: #7f8c8d; font-size: 12px; font-weight: bold; margin-bottom: 5px;">INGRESOS</div>
            <div style="color: {color_ingresos}; font-size: 24px; font-weight: bold;">
                ${total_ingresos:,.2f}
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        color_gastos = "#e74c3c"
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); 
                    padding: 20px; border-radius: 10px; text-align: center; 
                    border-left: 5px solid {color_gastos}; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
            <div style="color: #7f8c8d; font-size: 12px; font-weight: bold; margin-bottom: 5px;">GASTOS</div>
            <div style="color: {color_gastos}; font-size: 24px; font-weight: bold;">
                ${total_gastos:,.2f}
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        color_resultado = "#27ae60" if resultado >= 0 else "#e74c3c"
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); 
                    padding: 20px; border-radius: 10px; text-align: center; 
                    border-left: 5px solid {color_resultado}; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
            <div style="color: #7f8c8d; font-size: 12px; font-weight: bold; margin-bottom: 5px;">RESULTADO</div>
            <div style="color: {color_resultado}; font-size: 24px; font-weight: bold;">
                ${resultado:,.2f}
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        color_margen = "#27ae60" if margen_porcentaje >= 0 else "#e74c3c"
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); 
                    padding: 20px; border-radius: 10px; text-align: center; 
                    border-left: 5px solid {color_margen}; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
            <div style="color: #7f8c8d; font-size: 12px; font-weight: bold; margin-bottom: 5px;">MARGEN</div>
            <div style="color: {color_margen}; font-size: 24px; font-weight: bold;">
                {margen_porcentaje:.2f}%
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # SECCIÓN DE INGRESOS
    st.markdown("""
    <div style="background-color: #34495e; color: white; padding: 10px; border-radius: 5px; margin-bottom: 15px;">
        <h3 style="margin: 0; font-size: 18px;">VENTAS/INGRESOS</h3>
    </div>
    """, unsafe_allow_html=True)
    
    # Agrupar ingresos por categoría si hay datos
    if not df_ingresos.empty:
        # Aquí podrías agregar lógica para agrupar por tipo de ingreso si tienes esa información
        ingresos_df = pd.DataFrame({
            'Concepto': ['Ventas Salon'],
            'Monto': [total_ingresos]
        })
        
        for _, row in ingresos_df.iterrows():
            st.markdown(f"""
            <div style="padding: 8px 0; border-bottom: 1px solid #ecf0f1; display: flex; justify-content: space-between;">
                <span style="color: #2c3e50;">{row['Concepto']}</span>
                <span style="color: #2c3e50; font-weight: 500;">${row['Monto']:,.2f}</span>
            </div>
            """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style="padding: 8px 0; border-bottom: 1px solid #ecf0f1; display: flex; justify-content: space-between;">
            <span style="color: #2c3e50;">Sin ingresos registrados</span>
            <span style="color: #2c3e50; font-weight: 500;">$0.00</span>
        </div>
        """, unsafe_allow_html=True)
    
    # Total de ingresos
    st.markdown(f"""
    <div style="padding: 12px 0; margin-top: 10px; border-top: 2px solid #34495e; display: flex; justify-content: space-between;">
        <span style="color: #2c3e50; font-weight: bold; font-size: 16px;">TOTAL INGRESOS</span>
        <span style="color: #2c3e50; font-weight: bold; font-size: 16px;">${total_ingresos:,.2f}</span>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # SECCIÓN DE GASTOS
    st.markdown("""
    <div style="background-color: #34495e; color: white; padding: 10px; border-radius: 5px; margin-bottom: 15px;">
        <h3 style="margin: 0; font-size: 18px;">COMPRAS/EGRESOS</h3>
    </div>
    """, unsafe_allow_html=True)
    
    # Agrupar gastos por rubro con subtotales
    if 'rubro' in df_gastos.columns:
        gastos_agrupados = df_gastos.groupby('rubro')['total'].sum().sort_values(ascending=False)
        
        # Definir orden y categorías principales
        categorias_principales = {
            'ALIMENTOS': {'subcategorias': []},
            'BEBIDAS': {'subcategorias': []},
            'SUELDOS': {'subcategorias': ['Cargas Sociales']},
            'SERVICIOS': {'subcategorias': []},
            'ALQUILER': {'subcategorias': []},
            'INSUMOS Y DESCARTABLES': {'subcategorias': []},
            'IMPUESTOS': {'subcategorias': []},
            'MATERIALES': {'subcategorias': []}
        }
        
        # Procesar cada categoría principal
        for categoria, config in categorias_principales.items():
            # Buscar gastos de esta categoría
            gastos_categoria = gastos_agrupados[gastos_agrupados.index.str.contains(categoria, case=False, na=False)]
            
            if not gastos_categoria.empty:
                # Mostrar categoría principal
                total_categoria = gastos_categoria.sum()
                st.markdown(f"""
                <div style="padding: 8px 0; border-bottom: 1px solid #ecf0f1; display: flex; justify-content: space-between;">
                    <span style="color: #2c3e50; font-weight: 500;">{categoria.title()}</span>
                    <span style="color: #2c3e50; font-weight: 500;">${total_categoria:,.2f}</span>
                </div>
                """, unsafe_allow_html=True)
                
                # Mostrar subcategorías si existen
                for subcat in config['subcategorias']:
                    gastos_subcat = gastos_agrupados[gastos_agrupados.index.str.contains(subcat, case=False, na=False)]
                    if not gastos_subcat.empty:
                        for subcat_nombre, monto in gastos_subcat.items():
                            st.markdown(f"""
                            <div style="padding: 6px 0 6px 20px; border-bottom: 1px solid #ecf0f1; display: flex; justify-content: space-between; font-size: 14px;">
                                <span style="color: #7f8c8d;">└─ {subcat_nombre}</span>
                                <span style="color: #7f8c8d;">${monto:,.2f}</span>
                            </div>
                            """, unsafe_allow_html=True)
        
        # Mostrar otras categorías no incluidas en las principales
        otras_categorias = []
        for categoria, monto in gastos_agrupados.items():
            es_principal = any(cat in categoria.upper() for cat in categorias_principales.keys())
            if not es_principal:
                otras_categorias.append((categoria, monto))
        
        if otras_categorias:
            total_otras = sum(monto for _, monto in otras_categorias)
            st.markdown(f"""
            <div style="padding: 8px 0; border-bottom: 1px solid #ecf0f1; display: flex; justify-content: space-between;">
                <span style="color: #2c3e50; font-weight: 500;">Otros Gastos</span>
                <span style="color: #2c3e50; font-weight: 500;">${total_otras:,.2f}</span>
            </div>
            """, unsafe_allow_html=True)
            
            for categoria, monto in otras_categorias:
                st.markdown(f"""
                <div style="padding: 6px 0 6px 20px; border-bottom: 1px solid #ecf0f1; display: flex; justify-content: space-between; font-size: 14px;">
                    <span style="color: #7f8c8d;">└─ {categoria}</span>
                    <span style="color: #7f8c8d;">${monto:,.2f}</span>
                </div>
                """, unsafe_allow_html=True)
    else:
        # Si no hay rubro, mostrar total general
        st.markdown(f"""
        <div style="padding: 8px 0; border-bottom: 1px solid #ecf0f1; display: flex; justify-content: space-between;">
            <span style="color: #2c3e50;">Gastos Operativos</span>
            <span style="color: #2c3e50; font-weight: 500;">${total_gastos:,.2f}</span>
        </div>
        """, unsafe_allow_html=True)
    
    # Total de egresos
    st.markdown(f"""
    <div style="padding: 12px 0; margin-top: 10px; border-top: 2px solid #34495e; display: flex; justify-content: space-between;">
        <span style="color: #2c3e50; font-weight: bold; font-size: 16px;">TOTAL EGRESOS</span>
        <span style="color: #2c3e50; font-weight: bold; font-size: 16px;">${total_gastos:,.2f}</span>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # RESULTADO OPERATIVO
    color_resultado_final = "#27ae60" if resultado >= 0 else "#e74c3c"
    st.markdown(f"""
    <div style="background-color: {color_resultado_final}20; padding: 20px; border-radius: 10px; border-left: 5px solid {color_resultado_final}; margin-bottom: 30px;">
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <span style="color: {color_resultado_final}; font-weight: bold; font-size: 18px;">RESULTADO OPERATIVO</span>
            <span style="color: {color_resultado_final}; font-weight: bold; font-size: 24px;">${resultado:,.2f}</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # ANÁLISIS DE COMPOSICIÓN
    st.markdown("""
    <div style="background-color: #34495e; color: white; padding: 10px; border-radius: 5px; margin-bottom: 15px;">
        <h3 style="margin: 0; font-size: 18px;">ANÁLISIS DE COMPOSICIÓN</h3>
    </div>
    """, unsafe_allow_html=True)
    
    # Calcular porcentajes y comparar con benchmarks
    benchmarks = calcular_benchmarks_gastronomia()
    
    if 'rubro' in df_gastos.columns:
        gastos_por_rubro = df_gastos.groupby('rubro')['total'].sum()
        
        for rubro_key, benchmark in benchmarks.items():
            # Buscar gastos que coincidan con este rubro
            gastos_rubro = gastos_por_rubro[gastos_por_rubro.index.str.contains(rubro_key, case=False, na=False)]
            
            if not gastos_rubro.empty:
                total_rubro = gastos_rubro.sum()
                porcentaje_real = (total_rubro / total_ingresos * 100) if total_ingresos > 0 else 0
                
                # Determinar estado
                if porcentaje_real < benchmark['rango_min']:
                    estado = "BAJO"
                    icono = "⬇️"
                    color = "#3498db"
                elif porcentaje_real > benchmark['rango_max']:
                    estado = "ALTO"
                    icono = "⬆️"
                    color = "#e74c3c"
                else:
                    estado = "OK"
                    icono = "✅"
                    color = "#27ae60"
                
                st.markdown(f"""
                <div style="padding: 10px; margin-bottom: 10px; background-color: #f8f9fa; border-radius: 5px; border-left: 3px solid {color};">
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <span style="color: #2c3e50; font-weight: 500;">{icono} {rubro_key.replace('_', ' ').title()} sobre Ventas</span>
                        <span style="color: {color}; font-weight: bold;">{porcentaje_real:.2f}%</span>
                    </div>
                    <div style="font-size: 12px; color: #7f8c8d; margin-top: 5px;">
                        Ideal: {benchmark['rango_min']:.0f}%-{benchmark['rango_max']:.0f}% | Estado: {estado}
                    </div>
                </div>
                """, unsafe_allow_html=True)
    
    # Margen neto
    if margen_porcentaje >= 10:
        estado_margen = "EXCELENTE"
        icono_margen = "🟢"
        color_margen = "#27ae60"
    elif margen_porcentaje >= 5:
        estado_margen = "ACEPTABLE"
        icono_margen = "🟡"
        color_margen = "#f39c12"
    else:
        estado_margen = "CRÍTICO"
        icono_margen = "🔴"
        color_margen = "#e74c3c"
    
    st.markdown(f"""
    <div style="padding: 10px; margin-bottom: 10px; background-color: #f8f9fa; border-radius: 5px; border-left: 3px solid {color_margen};">
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <span style="color: #2c3e50; font-weight: 500;">{icono_margen} Margen Neto</span>
            <span style="color: {color_margen}; font-weight: bold;">{margen_porcentaje:.2f}%</span>
        </div>
        <div style="font-size: 12px; color: #7f8c8d; margin-top: 5px;">
            Ideal: 10%-15% | Estado: {estado_margen}
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Pie del informe
    st.markdown("---")
    st.markdown(f"""
    <div style="text-align: center; color: #7f8c8d; font-size: 12px; padding: 20px;">
        Informe generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | 
        Período: {nombre_mes} {anio_seleccionado} | 
        Sucursal: {sucursal_nombre}
    </div>
    """, unsafe_allow_html=True)
    
    # Boton de descarga Excel
    st.markdown("---")
    col_desc1, col_desc2, col_desc3 = st.columns([1, 2, 1])
    with col_desc2:
        excel_file = generar_excel_con_detalle(
            df_ingresos, df_gastos, total_ingresos, total_gastos, 
            resultado, margen_porcentaje, sucursal_nombre, 
            mes_seleccionado, anio_seleccionado
        )
        st.download_button(
            label="Descargar Estado de Resultados (Excel)",
            data=excel_file,
            file_name=f"Estado_Resultados_{sucursal_nombre}_{mes_seleccionado}_{anio_seleccionado}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
            key="download_analisis"
        )


def generar_pdf_estado_resultados(df_merged, df_ingresos, sucursal_nombre, mes_seleccionado,
                                   anio_seleccionado, total_ingresos, total_gastos):
    """
    Genera el Estado de Resultados Granular en formato PDF ejecutivo.
    Retorna bytes del PDF listo para descargar.
    """
    buffer = io.BytesIO()
    
    meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    nombre_mes = meses[mes_seleccionado]

    # ── Colores corporativos ──────────────────────────────────────────────────
    COLOR_HEADER    = colors.HexColor('#2c3e50')   # azul oscuro
    COLOR_SECCION   = colors.HexColor('#34495e')   # gris azulado
    COLOR_SUBTOTAL  = colors.HexColor('#7f8c8d')   # gris
    COLOR_VERDE     = colors.HexColor('#27ae60')
    COLOR_ROJO      = colors.HexColor('#e74c3c')
    COLOR_FILA_PAR  = colors.HexColor('#f8f9fa')
    COLOR_FILA_IMPAR= colors.white
    COLOR_SUBRUBRO  = colors.HexColor('#ecf0f1')

    # ── Estilos de párrafo ────────────────────────────────────────────────────
    styles = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle('titulo',
        fontSize=18, fontName='Helvetica-Bold',
        textColor=COLOR_HEADER, alignment=TA_CENTER, spaceAfter=4)

    estilo_subtitulo = ParagraphStyle('subtitulo',
        fontSize=13, fontName='Helvetica',
        textColor=COLOR_SECCION, alignment=TA_CENTER, spaceAfter=2)

    estilo_sucursal = ParagraphStyle('sucursal',
        fontSize=11, fontName='Helvetica',
        textColor=colors.HexColor('#7f8c8d'), alignment=TA_CENTER, spaceAfter=10)

    estilo_pie = ParagraphStyle('pie',
        fontSize=8, fontName='Helvetica',
        textColor=colors.HexColor('#95a5a6'), alignment=TA_CENTER)

    # ── Documento A4 ─────────────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
        title=f"Estado de Resultados - {sucursal_nombre} - {nombre_mes} {anio_seleccionado}",
        author="Sistema P&L"
    )

    story = []
    ancho_util = A4[0] - 3.6*cm   # ancho disponible

    # ── Función auxiliar: fila de item ────────────────────────────────────────
    def calcular_item(cod_inf):
        """Retorna (total_item, lista_subrubros) sin renderizar pantalla"""
        sub = df_merged[df_merged['cod_inf'] == cod_inf]
        if sub.empty:
            return 0, []
        total = sub['total'].sum()
        agrupados = sub.groupby('subrubro_norm')['total'].sum()
        detalles = [(srub, monto) for srub, monto in agrupados.items() if monto != 0]
        return total, detalles

    def fmt_pesos(valor):
        """Formatea un número como moneda argentina"""
        if valor < 0:
            return f"(${abs(valor):,.2f})"
        return f"${valor:,.2f}"

    def fmt_pct(valor, base):
        if base == 0:
            return "0.00%"
        return f"{valor/base*100:.2f}%"

    # ── Estilo base para tablas de items ──────────────────────────────────────
    def tabla_items_style(filas_datos, color_impar=COLOR_FILA_IMPAR, color_par=COLOR_FILA_PAR):
        cmds = [
            ('FONTNAME',    (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE',    (0,0), (-1,-1), 9),
            ('TEXTCOLOR',   (0,0), (-1,-1), COLOR_HEADER),
            ('ALIGN',       (1,0), (-1,-1), 'RIGHT'),
            ('LINEBELOW',   (0,0), (-1,-1), 0.3, colors.HexColor('#ecf0f1')),
            ('LEFTPADDING',  (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING',   (0,0), (-1,-1), 4),
            ('BOTTOMPADDING',(0,0), (-1,-1), 4),
        ]
        for i in range(filas_datos):
            bg = color_par if i % 2 == 0 else color_impar
            cmds.append(('BACKGROUND', (0,i), (-1,i), bg))
        return TableStyle(cmds)

    # ════════════════════════════════════════════════════════════════════════
    # CABECERA DEL INFORME
    # ════════════════════════════════════════════════════════════════════════
    story.append(Paragraph("ESTADO DE RESULTADOS GRANULAR", estilo_titulo))
    story.append(Paragraph(f"{nombre_mes.upper()} {anio_seleccionado}", estilo_subtitulo))
    story.append(Paragraph(sucursal_nombre.upper(), estilo_sucursal))
    story.append(HRFlowable(width="100%", thickness=2, color=COLOR_HEADER, spaceAfter=10))

    # ── KPIs resumen (4 celdas) ───────────────────────────────────────────────
    resultado_kpi  = total_ingresos - total_gastos
    margen_kpi     = (resultado_kpi / total_ingresos * 100) if total_ingresos > 0 else 0
    color_res_kpi  = COLOR_VERDE if resultado_kpi >= 0 else COLOR_ROJO
    color_mar_kpi  = COLOR_VERDE if margen_kpi >= 0 else COLOR_ROJO

    kpi_data = [[
        Paragraph(f'<font color="#7f8c8d" size="8"><b>INGRESOS</b></font><br/>'
                  f'<font color="{COLOR_VERDE.hexval()}" size="12"><b>{fmt_pesos(total_ingresos)}</b></font>', styles['Normal']),
        Paragraph(f'<font color="#7f8c8d" size="8"><b>GASTOS</b></font><br/>'
                  f'<font color="{COLOR_ROJO.hexval()}" size="12"><b>{fmt_pesos(total_gastos)}</b></font>', styles['Normal']),
        Paragraph(f'<font color="#7f8c8d" size="8"><b>RESULTADO</b></font><br/>'
                  f'<font color="{color_res_kpi.hexval()}" size="12"><b>{fmt_pesos(resultado_kpi)}</b></font>', styles['Normal']),
        Paragraph(f'<font color="#7f8c8d" size="8"><b>MARGEN</b></font><br/>'
                  f'<font color="{color_mar_kpi.hexval()}" size="12"><b>{margen_kpi:.2f}%</b></font>', styles['Normal']),
    ]]
    w4 = ancho_util / 4
    t_kpi = Table(kpi_data, colWidths=[w4]*4)
    t_kpi.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,-1), COLOR_FILA_PAR),
        ('BOX',          (0,0), (-1,-1), 0.5, colors.HexColor('#dee2e6')),
        ('INNERGRID',    (0,0), (-1,-1), 0.3, colors.HexColor('#dee2e6')),
        ('ALIGN',        (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',   (0,0), (-1,-1), 8),
        ('BOTTOMPADDING',(0,0), (-1,-1), 8),
        ('ROUNDEDCORNERS', [4]),
    ]))
    story.append(t_kpi)
    story.append(Spacer(1, 14))

    # ════════════════════════════════════════════════════════════════════════
    # SECCIÓN INGRESOS
    # ════════════════════════════════════════════════════════════════════════
    def header_seccion(texto):
        t = Table([[Paragraph(f'<font color="white"><b>{texto}</b></font>',
                              ParagraphStyle('sh', fontSize=11, fontName='Helvetica-Bold',
                                             textColor=colors.white))]],
                  colWidths=[ancho_util])
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,-1), COLOR_SECCION),
            ('TOPPADDING',    (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING',   (0,0), (-1,-1), 8),
        ]))
        return t

    story.append(header_seccion("VENTAS / INGRESOS"))
    story.append(Spacer(1, 4))

    # Filas de ingresos
    filas_ing = []
    COL_W = [ancho_util*0.55, ancho_util*0.25, ancho_util*0.20]

    if not df_ingresos.empty and 'categoria_id' in df_ingresos.columns:
        ing_otros = df_ingresos[df_ingresos['categoria_id'] != 2]['monto'].sum()
        ing_pya   = df_ingresos[df_ingresos['categoria_id'] == 2]['monto'].sum()
        if ing_otros > 0:
            filas_ing.append(["Ventas / Ingresos (otros medios)",
                               fmt_pesos(ing_otros),
                               fmt_pct(ing_otros, total_ingresos)])
        if ing_pya > 0:
            filas_ing.append(["Ventas Pedidos Ya",
                               fmt_pesos(ing_pya),
                               fmt_pct(ing_pya, total_ingresos)])
    elif not df_ingresos.empty:
        filas_ing.append(["Ventas / Ingresos",
                           fmt_pesos(total_ingresos), "100.00%"])

    if filas_ing:
        t = Table(filas_ing, colWidths=COL_W)
        t.setStyle(tabla_items_style(len(filas_ing)))
        story.append(t)

    # Total ingresos
    t_tot_ing = Table(
        [[Paragraph('<b>TOTAL INGRESOS</b>',
                    ParagraphStyle('ti', fontSize=10, fontName='Helvetica-Bold', textColor=COLOR_HEADER)),
          Paragraph(f'<b>{fmt_pesos(total_ingresos)}</b>',
                    ParagraphStyle('tiv', fontSize=10, fontName='Helvetica-Bold',
                                   textColor=COLOR_HEADER, alignment=TA_RIGHT)),
          ""]],
        colWidths=COL_W)
    t_tot_ing.setStyle(TableStyle([
        ('LINEABOVE',     (0,0), (-1,-1), 1.5, COLOR_SECCION),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING',   (0,0), (-1,-1), 6),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
        ('ALIGN',         (1,0), (2,0), 'RIGHT'),
    ]))
    story.append(t_tot_ing)
    story.append(Spacer(1, 12))

    # ════════════════════════════════════════════════════════════════════════
    # SECCIÓN EGRESOS
    # ════════════════════════════════════════════════════════════════════════
    story.append(header_seccion("COMPRAS / EGRESOS"))
    story.append(Spacer(1, 4))

    def agregar_grupo(items_def, label_subtotal):
        """
        items_def: lista de (cod_inf, nombre)
        Agrega las filas del grupo al story y retorna el subtotal.
        """
        filas = []
        subtotal = 0
        for cod_inf, nombre in items_def:
            total_item, detalles = calcular_item(cod_inf)
            if total_item == 0:
                continue
            pct = fmt_pct(total_item, total_gastos)
            filas.append([
                Paragraph(f'<b>{nombre}</b>',
                          ParagraphStyle('ni', fontSize=9, fontName='Helvetica-Bold',
                                         textColor=COLOR_HEADER)),
                Paragraph(f'<b>{fmt_pesos(total_item)}</b>',
                          ParagraphStyle('niv', fontSize=9, fontName='Helvetica-Bold',
                                         textColor=COLOR_HEADER, alignment=TA_RIGHT)),
                Paragraph(f'<b>{pct}</b>',
                          ParagraphStyle('nip', fontSize=9, fontName='Helvetica-Bold',
                                         textColor=COLOR_HEADER, alignment=TA_RIGHT)),
            ])
            for srub, monto in detalles:
                pct_s = fmt_pct(monto, total_gastos)
                filas.append([
                    Paragraph(f'  └─ {srub.title()}',
                              ParagraphStyle('sr', fontSize=8, fontName='Helvetica',
                                             textColor=colors.HexColor('#7f8c8d'))),
                    Paragraph(fmt_pesos(monto),
                              ParagraphStyle('srv', fontSize=8, fontName='Helvetica',
                                             textColor=colors.HexColor('#7f8c8d'), alignment=TA_RIGHT)),
                    Paragraph(pct_s,
                              ParagraphStyle('srp', fontSize=8, fontName='Helvetica',
                                             textColor=colors.HexColor('#95a5a6'), alignment=TA_RIGHT)),
                ])
            subtotal += total_item

        if not filas:
            return 0

        t = Table(filas, colWidths=COL_W)
        cmds = [
            ('ALIGN',        (1,0), (-1,-1), 'RIGHT'),
            ('LINEBELOW',    (0,0), (-1,-1), 0.3, colors.HexColor('#ecf0f1')),
            ('LEFTPADDING',  (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING',   (0,0), (-1,-1), 3),
            ('BOTTOMPADDING',(0,0), (-1,-1), 3),
        ]
        # Alternar colores fila principal vs subrubro
        i = 0
        for cod_inf, nombre in items_def:
            tot_item, dets = calcular_item(cod_inf)
            if tot_item == 0:
                continue
            cmds.append(('BACKGROUND', (0,i), (-1,i), COLOR_FILA_PAR))
            i += 1
            for _ in dets:
                cmds.append(('BACKGROUND', (0,i), (-1,i), COLOR_SUBRUBRO))
                i += 1
        t.setStyle(TableStyle(cmds))
        story.append(t)

        # Subtotal del grupo
        if subtotal > 0:
            pct_st = fmt_pct(subtotal, total_gastos)
            t_sub = Table(
                [[Paragraph(f'→ <b>{label_subtotal}</b>',
                             ParagraphStyle('lst', fontSize=9, fontName='Helvetica-Bold',
                                            textColor=COLOR_HEADER)),
                  Paragraph(f'<b>{fmt_pesos(subtotal)}</b>',
                             ParagraphStyle('lstv', fontSize=9, fontName='Helvetica-Bold',
                                            textColor=COLOR_HEADER, alignment=TA_RIGHT)),
                  Paragraph(f'<b>{pct_st}</b>',
                             ParagraphStyle('lstp', fontSize=9, fontName='Helvetica-Bold',
                                            textColor=COLOR_SUBTOTAL, alignment=TA_RIGHT))]],
                colWidths=COL_W)
            t_sub.setStyle(TableStyle([
                ('LINEABOVE',     (0,0), (-1,-1), 1, COLOR_SUBTOTAL),
                ('TOPPADDING',    (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ('LEFTPADDING',   (0,0), (-1,-1), 6),
                ('RIGHTPADDING',  (0,0), (-1,-1), 6),
                ('ALIGN',         (1,0), (2,0), 'RIGHT'),
            ]))
            story.append(t_sub)
        story.append(Spacer(1, 6))
        return subtotal

    # ── Los 5 grupos operativos ───────────────────────────────────────────────
    st1 = agregar_grupo(
        [(1, "Costo de Mercadería Consumida (CMC)"),
         (2, "Gastos de Personal"),
         (3, "Cargas sociales"),
         (4, "Sindicatos")],
        "Subtotal sueldos y CMC")

    st2 = agregar_grupo(
        [(5, "Alquiler"),
         (6, "Servicios")],
        "Subtotal alquileres y servicios")

    st3 = agregar_grupo(
        [(7, "Honorarios profesionales"),
         (8, "Publicidad")],
        "Subtotal de honorarios")

    st4 = agregar_grupo(
        [(9, "Materiales"),
         (10, "Mantenimiento"),
         (11, "Bienes de uso")],
        "Subtotal de mantenimiento y bienes de uso")

    st5 = agregar_grupo(
        [(12, "Royalties"),
         (18, "Servicios administrativos/logísticos")],
        "Subtotal de royalties y logística")

    # ── Total egresos operativos ──────────────────────────────────────────────
    total_op = st1 + st2 + st3 + st4 + st5
    saldo_imp_ret = total_ingresos - total_op

    t_teo = Table(
        [[Paragraph('<b>TOTAL DE EGRESOS</b>',
                    ParagraphStyle('teo', fontSize=10, fontName='Helvetica-Bold', textColor=COLOR_HEADER)),
          Paragraph(f'<b>{fmt_pesos(total_op)}</b>',
                    ParagraphStyle('teov', fontSize=10, fontName='Helvetica-Bold',
                                   textColor=COLOR_HEADER, alignment=TA_RIGHT)),
          ""]],
        colWidths=COL_W)
    t_teo.setStyle(TableStyle([
        ('LINEABOVE',     (0,0), (-1,-1), 1.5, COLOR_SECCION),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING',   (0,0), (-1,-1), 6),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
        ('ALIGN',         (1,0), (2,0), 'RIGHT'),
    ]))
    story.append(t_teo)

    # ── Saldo para impuestos y retiros ────────────────────────────────────────
    t_saldo = Table(
        [[Paragraph('<b>SALDO PARA IMPUESTOS Y RETIROS</b>',
                    ParagraphStyle('sal', fontSize=10, fontName='Helvetica-Bold', textColor=COLOR_VERDE)),
          Paragraph(f'<b>{fmt_pesos(saldo_imp_ret)}</b>',
                    ParagraphStyle('salv', fontSize=10, fontName='Helvetica-Bold',
                                   textColor=COLOR_VERDE, alignment=TA_RIGHT)),
          ""]],
        colWidths=COL_W)
    t_saldo.setStyle(TableStyle([
        ('LINEABOVE',     (0,0), (-1,-1), 1, COLOR_VERDE),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING',   (0,0), (-1,-1), 6),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
        ('ALIGN',         (1,0), (2,0), 'RIGHT'),
    ]))
    story.append(t_saldo)
    story.append(Spacer(1, 8))

    # ── Grupo 6: Retiros ──────────────────────────────────────────────────────
    tot_ret = agregar_grupo([(19, "Retiros")], "Subtotal de retiros")

    # ── Grupo 7: Impuestos y bancarios ────────────────────────────────────────
    st7 = agregar_grupo(
        [(13, "Gastos y comisiones bancarias/tarjetas"),
         (14, "Impuestos municipales"),
         (15, "Impuestos provinciales"),
         (16, "Impuestos nacionales"),
         (17, "Iva a pagar")],
        "Subtotal de impuestos")

    # ── TOTAL EGRESOS FINAL ───────────────────────────────────────────────────
    total_final = total_op + tot_ret + st7
    t_tef = Table(
        [[Paragraph('<b>TOTAL EGRESOS</b>',
                    ParagraphStyle('tef', fontSize=11, fontName='Helvetica-Bold', textColor=COLOR_HEADER)),
          Paragraph(f'<b>{fmt_pesos(total_final)}</b>',
                    ParagraphStyle('tefv', fontSize=11, fontName='Helvetica-Bold',
                                   textColor=COLOR_HEADER, alignment=TA_RIGHT)),
          ""]],
        colWidths=COL_W)
    t_tef.setStyle(TableStyle([
        ('LINEABOVE',     (0,0), (-1,-1), 2, COLOR_SECCION),
        ('TOPPADDING',    (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING',   (0,0), (-1,-1), 6),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
        ('ALIGN',         (1,0), (2,0), 'RIGHT'),
    ]))
    story.append(t_tef)
    story.append(Spacer(1, 12))

    # ════════════════════════════════════════════════════════════════════════
    # RESULTADO OPERATIVO FINAL
    # ════════════════════════════════════════════════════════════════════════
    resultado_final = total_ingresos - total_final
    pct_res_final   = (resultado_final / total_ingresos * 100) if total_ingresos > 0 else 0
    color_res_final = COLOR_VERDE if resultado_final >= 0 else COLOR_ROJO
    texto_signo     = "GANANCIA" if resultado_final >= 0 else "PÉRDIDA"

    t_res = Table(
        [[Paragraph(f'<b>RESULTADO OPERATIVO ESTIMADO</b>',
                    ParagraphStyle('rof', fontSize=12, fontName='Helvetica-Bold',
                                   textColor=color_res_final)),
          Paragraph(f'<b>{fmt_pesos(resultado_final)}</b>',
                    ParagraphStyle('rofv', fontSize=14, fontName='Helvetica-Bold',
                                   textColor=color_res_final, alignment=TA_RIGHT)),
          Paragraph(f'<b>({pct_res_final:.2f}%)</b>',
                    ParagraphStyle('rofp', fontSize=11, fontName='Helvetica-Bold',
                                   textColor=color_res_final, alignment=TA_RIGHT))]],
        colWidths=COL_W)
    t_res.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), colors.HexColor(
            '#d5f5e3' if resultado_final >= 0 else '#fadbd8')),
        ('LINEABOVE',     (0,0), (-1,-1), 2, color_res_final),
        ('LINEBEFORE',    (0,0), (0,-1), 4, color_res_final),
        ('TOPPADDING',    (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('LEFTPADDING',   (0,0), (-1,-1), 10),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
        ('ALIGN',         (1,0), (2,0), 'RIGHT'),
        ('ROUNDEDCORNERS', [4]),
    ]))
    story.append(KeepTogether(t_res))

    # ── Pie de página ─────────────────────────────────────────────────────────
    story.append(Spacer(1, 16))
    story.append(HRFlowable(width="100%", thickness=0.5, color=COLOR_SUBTOTAL))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"Informe generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}  |  "
        f"Período: {nombre_mes} {anio_seleccionado}  |  Sucursal: {sucursal_nombre}",
        estilo_pie))

    # ── Construir PDF ─────────────────────────────────────────────────────────
    doc.build(story)
    buffer.seek(0)
    return buffer.read()


def mostrar_estado_resultados_granular(supabase, sucursales, mes_seleccionado, anio_seleccionado, sucursal_seleccionada):
    """
    Genera el Estado de Resultados Granular con estructura jerárquica y diseño profesional
    """
    # Header con botón de refrescar
    col_header, col_refresh = st.columns([4, 1])
    
    with col_header:
        st.subheader("📊 Estado de Resultados Granular")
    
    with col_refresh:
        if st.button("🔄 Refrescar", key="refresh_granular", width="stretch"):
            limpiar_cache_pl_simples()
            st.success("✅ Datos actualizados")
            st.rerun()
    
    sucursal_id = sucursal_seleccionada['id'] if sucursal_seleccionada else None
    sucursal_nombre = sucursal_seleccionada['nombre'] if sucursal_seleccionada else "Todas las sucursales"
    
    # Obtener datos
    with st.spinner("Cargando datos..."):
        df_gastos = obtener_gastos_db(supabase, mes_seleccionado, anio_seleccionado, sucursal_id)
        df_ingresos = obtener_ingresos_mensuales(supabase, mes_seleccionado, anio_seleccionado, sucursal_id)
    
    if df_gastos.empty:
        st.warning(f"⚠️ No hay gastos registrados para **{sucursal_nombre}** en **{mes_seleccionado}/{anio_seleccionado}**.")
        return
    
    # Obtener mapeo de BD
    try:
        result_mapeo = supabase.table("mapeo_estado_resultado_granular").select("*").execute()
        if result_mapeo.data:
            df_mapeo = pd.DataFrame(result_mapeo.data)
            df_mapeo['subrubro'] = df_mapeo['subrubro'].str.upper().str.strip()
        else:
            st.warning("⚠️ No hay mapeo configurado")
            return
    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
        return
    
    # Normalizar y hacer merge
    columna_subrubro = 'subrubro' if 'subrubro' in df_gastos.columns else 'rubro'
    df_gastos['subrubro_norm'] = df_gastos[columna_subrubro].str.upper().str.strip()
    
    df_merged = df_gastos.merge(
        df_mapeo[['cod_inf', 'item', 'subrubro']],
        left_on='subrubro_norm',
        right_on='subrubro',
        how='left'
    )
    
    # Calcular totales
    total_ingresos = df_ingresos['monto'].sum() if not df_ingresos.empty else 0
    total_gastos = df_gastos['total'].sum()
    resultado = total_ingresos - total_gastos
    margen_porcentaje = (resultado / total_ingresos * 100) if total_ingresos > 0 else 0
    
    # ==================================================================================
    # HEADER DEL INFORME
    # ==================================================================================
    st.markdown("---")
    
    meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    nombre_mes = meses[mes_seleccionado]
    
    st.markdown(f"""
    <div style="text-align: center; padding: 20px; background-color: #f8f9fa; border-radius: 10px; margin-bottom: 30px;">
        <h1 style="color: #2c3e50; margin-bottom: 10px;">ESTADO DE RESULTADOS GRANULAR</h1>
        <h2 style="color: #34495e; margin-bottom: 5px;">{nombre_mes.upper()} {anio_seleccionado}</h2>
        <h3 style="color: #7f8c8d; font-weight: normal;">{sucursal_nombre.upper()}</h3>
    </div>
    """, unsafe_allow_html=True)
    
    # ==================================================================================
    # KPIs EN TARJETAS
    # ==================================================================================
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        color_ingresos = "#27ae60" if total_ingresos > 0 else "#95a5a6"
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); 
                    padding: 20px; border-radius: 10px; text-align: center; 
                    border-left: 5px solid {color_ingresos}; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
            <div style="color: #7f8c8d; font-size: 12px; font-weight: bold; margin-bottom: 5px;">INGRESOS</div>
            <div style="color: {color_ingresos}; font-size: 24px; font-weight: bold;">
                ${total_ingresos:,.2f}
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); 
                    padding: 20px; border-radius: 10px; text-align: center; 
                    border-left: 5px solid #e74c3c; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
            <div style="color: #7f8c8d; font-size: 12px; font-weight: bold; margin-bottom: 5px;">GASTOS</div>
            <div style="color: #e74c3c; font-size: 24px; font-weight: bold;">
                ${total_gastos:,.2f}
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        color_resultado = "#27ae60" if resultado >= 0 else "#e74c3c"
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); 
                    padding: 20px; border-radius: 10px; text-align: center; 
                    border-left: 5px solid {color_resultado}; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
            <div style="color: #7f8c8d; font-size: 12px; font-weight: bold; margin-bottom: 5px;">RESULTADO</div>
            <div style="color: {color_resultado}; font-size: 24px; font-weight: bold;">
                ${resultado:,.2f}
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        color_margen = "#27ae60" if margen_porcentaje >= 0 else "#e74c3c"
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); 
                    padding: 20px; border-radius: 10px; text-align: center; 
                    border-left: 5px solid {color_margen}; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
            <div style="color: #7f8c8d; font-size: 12px; font-weight: bold; margin-bottom: 5px;">MARGEN</div>
            <div style="color: {color_margen}; font-size: 24px; font-weight: bold;">
                {margen_porcentaje:.2f}%
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # ==================================================================================
    # FUNCIÓN AUXILIAR PARA MOSTRAR ITEMS
    # ==================================================================================
    def mostrar_item(cod_inf, nombre_item, df_merged):
        """Muestra un item con sus subrubros"""
        # Filtrar subrubros de este item
        subrubros_item = df_merged[df_merged['cod_inf'] == cod_inf]
        
        if subrubros_item.empty:
            return 0
        
        # Calcular total del item
        total_item = subrubros_item['total'].sum()
        
        # Calcular porcentaje sobre total gastos
        porcentaje_item = (total_item / total_gastos * 100) if total_gastos > 0 else 0
        
        # Mostrar item principal con porcentaje
        st.markdown(f"""
        <div style="padding: 8px 0; border-bottom: 1px solid #ecf0f1; display: flex; justify-content: space-between; align-items: center;">
            <span style="color: #2c3e50; font-weight: 500;">{nombre_item}</span>
            <div style="display: flex; align-items: center; gap: 15px;">
                <span style="color: #2c3e50; font-weight: 500;">${total_item:,.2f}</span>
                <span style="color: #7f8c8d; font-size: 13px; min-width: 60px; text-align: right;">{porcentaje_item:.2f}%</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Agrupar subrubros y mostrar con porcentajes
        subrubros_agrupados = subrubros_item.groupby('subrubro_norm')['total'].sum()
        for subrubro, monto in subrubros_agrupados.items():
            if monto > 0:
                porcentaje_subrubro = (monto / total_gastos * 100) if total_gastos > 0 else 0
                st.markdown(f"""
                <div style="padding: 6px 0 6px 20px; border-bottom: 1px solid #ecf0f1; display: flex; justify-content: space-between; align-items: center; font-size: 14px;">
                    <span style="color: #7f8c8d;">└─ {subrubro.title()}</span>
                    <div style="display: flex; align-items: center; gap: 15px;">
                        <span style="color: #7f8c8d;">${monto:,.2f}</span>
                        <span style="color: #95a5a6; font-size: 12px; min-width: 60px; text-align: right;">{porcentaje_subrubro:.2f}%</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)
        
        return total_item
    
    def mostrar_subtotal(texto, monto):
        """Muestra un subtotal"""
        porcentaje = (monto / total_gastos * 100) if total_gastos > 0 else 0
        st.markdown(f"""
        <div style="padding: 10px 0; margin-top: 5px; border-top: 2px solid #95a5a6; display: flex; justify-content: space-between; align-items: center;">
            <span style="color: #2c3e50; font-weight: bold; font-size: 15px;">→ {texto}</span>
            <div style="display: flex; align-items: center; gap: 15px;">
                <span style="color: #2c3e50; font-weight: bold; font-size: 15px;">${monto:,.2f}</span>
                <span style="color: #7f8c8d; font-size: 13px; min-width: 60px; text-align: right;">{porcentaje:.2f}%</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    # ==================================================================================
    # ==================================================================================
    # SECCIÓN DE INGRESOS - CON SEPARACIÓN DE PEDIDOS YA (POR CATEGORÍA)
    # ==================================================================================
    st.markdown("""
    <div style="background-color: #34495e; color: white; padding: 10px; border-radius: 5px; margin-bottom: 15px;">
        <h3 style="margin: 0; font-size: 18px;">VENTAS/INGRESOS</h3>
    </div>
    """, unsafe_allow_html=True)
    
    if not df_ingresos.empty:
        # Verificar si existe la columna categoria_id
        if 'categoria_id' in df_ingresos.columns:
            # Separar ingresos de Pedidos Ya (categoria_id = 2) del resto
            ingresos_pedidos_ya = df_ingresos[df_ingresos['categoria_id'] == 2]['monto'].sum()
            ingresos_otros_medios = df_ingresos[df_ingresos['categoria_id'] != 2]['monto'].sum()
            
            # Calcular porcentajes sobre total ingresos
            porcentaje_otros = (ingresos_otros_medios / total_ingresos * 100) if total_ingresos > 0 else 0
            porcentaje_pedidos_ya = (ingresos_pedidos_ya / total_ingresos * 100) if total_ingresos > 0 else 0
            
            # Mostrar Ventas/Ingresos (otros medios) - Solo si hay monto
            if ingresos_otros_medios > 0:
                st.markdown(f"""
                <div style="padding: 8px 0; border-bottom: 1px solid #ecf0f1; display: flex; justify-content: space-between; align-items: center;">
                    <span style="color: #2c3e50;">Ventas/Ingresos (otros medios)</span>
                    <div style="display: flex; align-items: center; gap: 15px;">
                        <span style="color: #2c3e50; font-weight: 500;">${ingresos_otros_medios:,.2f}</span>
                        <span style="color: #7f8c8d; font-size: 13px; min-width: 60px; text-align: right;">{porcentaje_otros:.2f}%</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            
            # Mostrar Ventas Pedidos Ya - Solo si hay monto
            if ingresos_pedidos_ya > 0:
                st.markdown(f"""
                <div style="padding: 8px 0; border-bottom: 1px solid #ecf0f1; display: flex; justify-content: space-between; align-items: center;">
                    <span style="color: #2c3e50;">Ventas Pedidos Ya</span>
                    <div style="display: flex; align-items: center; gap: 15px;">
                        <span style="color: #2c3e50; font-weight: 500;">${ingresos_pedidos_ya:,.2f}</span>
                        <span style="color: #7f8c8d; font-size: 13px; min-width: 60px; text-align: right;">{porcentaje_pedidos_ya:.2f}%</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            
        else:
            # Si no existe categoria_id, mostrar como antes (sin separación)
            st.markdown(f"""
            <div style="padding: 8px 0; border-bottom: 1px solid #ecf0f1; display: flex; justify-content: space-between; align-items: center;">
                <span style="color: #2c3e50;">Ventas/Ingresos</span>
                <div style="display: flex; align-items: center; gap: 15px;">
                    <span style="color: #2c3e50; font-weight: 500;">${total_ingresos:,.2f}</span>
                    <span style="color: #7f8c8d; font-size: 13px; min-width: 60px; text-align: right;">100.00%</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
    else:
        # Sin datos
        st.markdown("""
        <div style="padding: 8px 0; border-bottom: 1px solid #ecf0f1; display: flex; justify-content: space-between;">
            <span style="color: #2c3e50;">Ventas/Ingresos</span>
            <span style="color: #2c3e50; font-weight: 500;">$0.00</span>
        </div>
        """, unsafe_allow_html=True)
    
    # Total de ingresos
    st.markdown(f"""
    <div style="padding: 12px 0; margin-top: 10px; border-top: 2px solid #34495e; display: flex; justify-content: space-between;">
        <span style="color: #2c3e50; font-weight: bold; font-size: 16px;">TOTAL INGRESOS</span>
        <span style="color: #2c3e50; font-weight: bold; font-size: 16px;">${total_ingresos:,.2f}</span>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)

    
    # ==================================================================================
    # SECCIÓN DE GASTOS CON ESTRUCTURA JERÁRQUICA
    # ==================================================================================
    st.markdown("""
    <div style="background-color: #34495e; color: white; padding: 10px; border-radius: 5px; margin-bottom: 15px;">
        <h3 style="margin: 0; font-size: 18px;">COMPRAS/EGRESOS</h3>
    </div>
    """, unsafe_allow_html=True)
    
    # ==================================================================================
    # GRUPO 1: CMC + Sueldos + Cargas + Sindicatos
    # ==================================================================================
    subtotal_grupo1 = 0
    subtotal_grupo1 += mostrar_item(1, "Costo de Mercadería Consumida (CMC)", df_merged)
    subtotal_grupo1 += mostrar_item(2, "Gastos de Personal", df_merged)
    subtotal_grupo1 += mostrar_item(3, "Cargas sociales", df_merged)
    subtotal_grupo1 += mostrar_item(4, "Sindicatos", df_merged)
    
    if subtotal_grupo1 > 0:
        mostrar_subtotal("Subtotal sueldos y CMC", subtotal_grupo1)
        st.markdown("<br>", unsafe_allow_html=True)
    
    # ==================================================================================
    # GRUPO 2: Alquiler + Servicios
    # ==================================================================================
    subtotal_grupo2 = 0
    subtotal_grupo2 += mostrar_item(5, "Alquiler", df_merged)
    subtotal_grupo2 += mostrar_item(6, "Servicios", df_merged)
    
    if subtotal_grupo2 > 0:
        mostrar_subtotal("Subtotal alquileres y servicios", subtotal_grupo2)
        st.markdown("<br>", unsafe_allow_html=True)
    
    # ==================================================================================
    # GRUPO 3: Honorarios + Publicidad
    # ==================================================================================
    subtotal_grupo3 = 0
    subtotal_grupo3 += mostrar_item(7, "Honorarios profesionales", df_merged)
    subtotal_grupo3 += mostrar_item(8, "Publicidad", df_merged)
    
    if subtotal_grupo3 > 0:
        mostrar_subtotal("Subtotal de honorarios", subtotal_grupo3)
        st.markdown("<br>", unsafe_allow_html=True)
    
    # ==================================================================================
    # GRUPO 4: Materiales + Mantenimiento + Bienes de uso
    # ==================================================================================
    subtotal_grupo4 = 0
    subtotal_grupo4 += mostrar_item(9, "Materiales", df_merged)
    subtotal_grupo4 += mostrar_item(10, "Mantenimiento", df_merged)
    subtotal_grupo4 += mostrar_item(11, "Bienes de uso", df_merged)
    
    if subtotal_grupo4 > 0:
        mostrar_subtotal("Subtotal de mantenimiento y bienes de uso", subtotal_grupo4)
        st.markdown("<br>", unsafe_allow_html=True)
    
    # ==================================================================================
    # GRUPO 5: Royalties + Servicios admin
    # ==================================================================================
    subtotal_grupo5 = 0
    subtotal_grupo5 += mostrar_item(12, "Royalties", df_merged)
    subtotal_grupo5 += mostrar_item(18, "Servicios administrativos/logísticos", df_merged)
    
    if subtotal_grupo5 > 0:
        mostrar_subtotal("Subtotal de royalties y logística", subtotal_grupo5)
        st.markdown("<br>", unsafe_allow_html=True)
    
    # Total de egresos operativos
    total_egresos_operativos = subtotal_grupo1 + subtotal_grupo2 + subtotal_grupo3 + subtotal_grupo4 + subtotal_grupo5
    
    st.markdown(f"""
    <div style="padding: 12px 0; margin-top: 10px; border-top: 2px solid #34495e; display: flex; justify-content: space-between;">
        <span style="color: #2c3e50; font-weight: bold; font-size: 16px;">TOTAL DE EGRESOS</span>
        <span style="color: #2c3e50; font-weight: bold; font-size: 16px;">${total_egresos_operativos:,.2f}</span>
    </div>
    """, unsafe_allow_html=True)
    
    saldo_impuestos_retiros = total_ingresos - total_egresos_operativos
    
    st.markdown(f"""
    <div style="padding: 12px 0; border-top: 2px solid #27ae60; display: flex; justify-content: space-between;">
        <span style="color: #27ae60; font-weight: bold; font-size: 16px;">SALDO PARA IMPUESTOS Y RETIROS</span>
        <span style="color: #27ae60; font-weight: bold; font-size: 16px;">${saldo_impuestos_retiros:,.2f}</span>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # ==================================================================================
    # GRUPO 6: Retiros
    # ==================================================================================
    total_retiros = mostrar_item(19, "Retiros", df_merged)
    
    if total_retiros > 0:
        st.markdown("<br>", unsafe_allow_html=True)
    
    # ==================================================================================
    # GRUPO 7: Gastos bancarios + Impuestos
    # ==================================================================================
    subtotal_impuestos = 0
    subtotal_impuestos += mostrar_item(13, "Gastos y comisiones bancarias/tarjetas", df_merged)
    subtotal_impuestos += mostrar_item(14, "Impuestos municipales", df_merged)
    subtotal_impuestos += mostrar_item(15, "Impuestos provinciales", df_merged)
    subtotal_impuestos += mostrar_item(16, "Impuestos nacionales", df_merged)
    subtotal_impuestos += mostrar_item(17, "Iva a pagar", df_merged)
    
    if subtotal_impuestos > 0:
        mostrar_subtotal("Subtotal de impuestos", subtotal_impuestos)
    
    if total_retiros > 0:
        mostrar_subtotal("Subtotal de retiros", total_retiros)
    
    # ==================================================================================
    # TOTAL EGRESOS FINAL
    # ==================================================================================
    total_egresos_final = total_egresos_operativos + total_retiros + subtotal_impuestos
    
    st.markdown(f"""
    <div style="padding: 12px 0; margin-top: 10px; border-top: 2px solid #34495e; display: flex; justify-content: space-between;">
        <span style="color: #2c3e50; font-weight: bold; font-size: 16px;">TOTAL EGRESOS</span>
        <span style="color: #2c3e50; font-weight: bold; font-size: 16px;">${total_egresos_final:,.2f}</span>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # ==================================================================================
    # RESULTADO OPERATIVO
    # ==================================================================================
    resultado_final = total_ingresos - total_egresos_final
    porcentaje_resultado = (resultado_final / total_ingresos * 100) if total_ingresos > 0 else 0
    color_resultado_final = "#27ae60" if resultado_final >= 0 else "#e74c3c"
    
    st.markdown(f"""
    <div style="background-color: {color_resultado_final}20; padding: 20px; border-radius: 10px; border-left: 5px solid {color_resultado_final}; margin-bottom: 30px;">
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <span style="color: {color_resultado_final}; font-weight: bold; font-size: 18px;">RESULTADO OPERATIVO ESTIMADO</span>
            <div style="display: flex; align-items: center; gap: 20px;">
                <span style="color: {color_resultado_final}; font-weight: bold; font-size: 24px;">${resultado_final:,.2f}</span>
                <span style="color: {color_resultado_final}; font-weight: bold; font-size: 18px;">({porcentaje_resultado:.2f}%)</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Botón de descarga PDF ────────────────────────────────────────────────
    st.markdown("---")
    col_pdf1, col_pdf2, col_pdf3 = st.columns([1, 2, 1])
    with col_pdf2:
        meses_nombres = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        with st.spinner("Preparando PDF..."):
            pdf_bytes = generar_pdf_estado_resultados(
                df_merged, df_ingresos, sucursal_nombre,
                mes_seleccionado, anio_seleccionado,
                total_ingresos, total_gastos
            )
        nombre_pdf = f"Estado_Resultados_{sucursal_nombre}_{meses_nombres[mes_seleccionado]}_{anio_seleccionado}.pdf"
        st.download_button(
            label="📄 Descargar Estado de Resultados (PDF)",
            data=pdf_bytes,
            file_name=nombre_pdf,
            mime="application/pdf",
            width="stretch",
            key="download_granular_pdf"
        )


def mostrar_tab_evolucion(supabase, sucursales, mes_seleccionado, anio_seleccionado, sucursal_seleccionada):
    """
    Tab de evolución histórica
    """
    st.subheader("📈 Evolución Histórica")
    
    if not sucursal_seleccionada:
        st.warning("⚠️ Por favor, selecciona una sucursal para ver la evolución histórica.")
        return
    
    sucursal_id = sucursal_seleccionada['id']
    sucursal_nombre = sucursal_seleccionada['nombre']
    
    # Obtener datos históricos
    with st.spinner("Cargando datos históricos..."):
        df_evolucion = obtener_evolucion_historica(supabase, sucursal_id)
    
    if df_evolucion.empty:
        st.warning(f"⚠️ No hay datos históricos para **{sucursal_nombre}**.")
        return
    
    # Convertir período a formato legible
    df_evolucion['periodo_str'] = df_evolucion['periodo'].dt.strftime('%m/%Y')
    
    # Gráfico de evolución
    st.markdown("### 📊 Evolución de Ingresos vs Gastos")
    
    import plotly.express as px
    import plotly.graph_objects as go
    
    # Crear gráfico
    fig = go.Figure()
    
    # Agregar ingresos
    fig.add_trace(go.Scatter(
        x=df_evolucion['periodo_str'],
        y=df_evolucion['total_ingresos'],
        mode='lines+markers',
        name='Ingresos',
        line=dict(color='#27ae60', width=3),
        hovertemplate='<b>%{x}</b><br>Ingresos: $%{y:,.2f}<extra></extra>'
    ))
    
    # Agregar gastos
    fig.add_trace(go.Scatter(
        x=df_evolucion['periodo_str'],
        y=df_evolucion['total_gastos'],
        mode='lines+markers',
        name='Gastos',
        line=dict(color='#e74c3c', width=3),
        hovertemplate='<b>%{x}</b><br>Gastos: $%{y:,.2f}<extra></extra>'
    ))
    
    # Agregar resultado
    fig.add_trace(go.Scatter(
        x=df_evolucion['periodo_str'],
        y=df_evolucion['resultado'],
        mode='lines+markers',
        name='Resultado',
        line=dict(color='#3498db', width=2, dash='dash'),
        hovertemplate='<b>%{x}</b><br>Resultado: $%{y:,.2f}<extra></extra>'
    ))
    
    fig.update_layout(
        title=f"Evolución Financiera - {sucursal_nombre}",
        xaxis_title="Período",
        yaxis_title="Monto ($)",
        hovermode='x unified',
        template='plotly_white',
        height=500
    )
    
    st.plotly_chart(fig, width="stretch")
    
    # Tabla de datos
    st.markdown("### 📋 Datos Detallados")
    
    # Formatear tabla para mejor visualización
    df_display = df_evolucion[['periodo_str', 'total_ingresos', 'total_gastos', 'resultado', 'margen']].copy()
    df_display.columns = ['Período', 'Ingresos', 'Gastos', 'Resultado', 'Margen %']
    
    # Formatear columnas monetarias
    for col in ['Ingresos', 'Gastos', 'Resultado']:
        df_display[col] = df_display[col].apply(lambda x: f"${x:,.2f}")
    
    df_display['Margen %'] = df_display['Margen %'].apply(lambda x: f"{x:.2f}%")
    
    st.dataframe(df_display, hide_index=True, width="stretch")


def main(supabase):
    """
    Función principal de la aplicación
    Recibe: supabase (cliente de conexión)
    """
    st.subheader("💰 P&L Simples - Profit & Loss")
    st.markdown("---")
    
    # --- ELIMINADO: La lógica de buscar supabase en session_state ---
    # Ya que ahora lo recibimos directamente como argumento
    
    # Obtener sucursales
    try:
        result = supabase.table("sucursales").select("*").execute()
        sucursales = result.data if result.data else []
        
        # Agregar opción "Todas las sucursales"
        sucursales_con_todas = [{'id': None, 'nombre': 'Todas las sucursales'}] + sucursales
    except Exception as e:
        st.error(f"❌ Error obteniendo sucursales: {str(e)}")
        return
    
    # --- MODIFICADO: Layout para que funcione dentro de cajas_diarias ---
    # Usamos st.columns en lugar de st.sidebar para no saturar la barra lateral principal
    
    col_filtros, col_kpis = st.columns([1, 3])
    
    with col_filtros:
        st.markdown("### 🔍 Filtros")
        
        # Selector de sucursal
        sucursal_opciones = {s['nombre']: s for s in sucursales_con_todas}
        sucursal_seleccionada_nombre = st.selectbox(
            "Sucursal",
            options=list(sucursal_opciones.keys()),
            index=0,
            key="pl_sucursal_select" # Agregamos key única para evitar conflictos
        )
        sucursal_seleccionada = sucursal_opciones[sucursal_seleccionada_nombre]
        
        # Selectores de mes y año
        mes_actual = datetime.now().month
        anio_actual = datetime.now().year
        
        c1, c2 = st.columns(2)
        with c1:
            mes_seleccionado = st.selectbox(
                "Mes",
                options=list(range(1, 13)),
                format_func=lambda x: ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                                      'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][x],
                index=mes_actual - 1,
                key="pl_mes_select"
            )
        
        with c2:
            anio_seleccionado = st.selectbox(
                "Año",
                options=list(range(2023, anio_actual + 1)),
                index=anio_actual - 2023,
                key="pl_anio_select"
            )
    
    # Tabs principales
    tab1, tab2, tab3, tab4 = st.tabs(["📊 Análisis del Período", "📊 Estado de Resultado Granular", "📁 Importar Gastos", "📈 Evolución Histórica"])
    
    with tab1:
        mostrar_tab_analisis(supabase, sucursales, mes_seleccionado, anio_seleccionado, sucursal_seleccionada)
    
    with tab2:
        mostrar_estado_resultados_granular(supabase, sucursales, mes_seleccionado, anio_seleccionado, sucursal_seleccionada)
    
    with tab3:
        mostrar_tab_importacion(supabase, sucursales, mes_seleccionado, anio_seleccionado, sucursal_seleccionada)
    
    with tab4:
        mostrar_tab_evolucion(supabase, sucursales, mes_seleccionado, anio_seleccionado, sucursal_seleccionada)


if __name__ == "__main__":
    main()
